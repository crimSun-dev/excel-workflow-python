"""Report Giro master-balance update workflow.

Replaces the operator's manual Ctrl+F reconciliation: for every account in the
Giro master workbook, look up that month's IDR balance in the monthly giro
source workbook and write it into `SALDO UPDATE`. The original five-column
master (`JENIS`, `NO REK`, `PRODUK`, `OPEN DATE`, `SALDO`) is the normal input:
if that write column is missing, it is appended. A leftover month-named column
such as `SALDO JUNI` is still accepted when present.

This workflow does not fit the shared ingest -> aggregate -> export template and
so overrides `execute()` (the precedent set by Qlola):

* the raw/source input is an Excel workbook, not pipe-delimited text;
* the output is the *master workbook itself*, updated cell-by-cell with
  openpyxl and saved to the configured output path, so historical columns,
  styling, formulas, and unrelated sheets all survive for `.xlsx` masters. A
  pandas `to_excel` round-trip would destroy every one of them. Legacy `.xls`
  masters are converted into an openpyxl workbook first (openpyxl cannot
  edit BIFF8 in place) and the result is always saved as `.xlsx`;
* there is no Summary_Report pivot.

Business rules (see the `workflow-processing` spec):
    * `JENIS = Deposito` rows are never updated (case-insensitive).
    * Master accounts absent from the monthly source are left blank - the
      `UNMAPPED` sentinel must never land in a numeric balance cell.
    * The historical `SALDO` column is never written to. Missing
      `SALDO UPDATE` is created; it is not filled by overwriting `SALDO`.
"""

from __future__ import annotations

import re
from pathlib import Path
from time import perf_counter

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..enrichment import canonicalize_join_id, normalize_header, resolve_alias_column
from ..exporter import PLAIN_INTEGER_FORMAT, ExportError
from ..ingestion import EMPTY_INPUT_MESSAGE, promote_header_row
from ..schemas import ProcessingConfig
from ..xls_support import open_excel
from .base import (
    WorkflowDefinition,
    WorkflowId,
    WorkflowRunResult,
    WorkflowStrategy,
    WorkflowValidationError,
)

# Ordered alias registries (first match wins), mirroring `enrichment.py`. Both
# sides are trimmed and upper-cased before comparison, so `no rek` matches
# `NO REK` while the spaced/underscored spellings stay distinct entries.
ACCOUNT_ALIASES = [
    "NO REK",
    "NO_REK",
    "NOREK",
    "NO REKENING",
    "NO_REKENING",
    "NOMOR REKENING",
    "REKENING",
]
JENIS_ALIASES = ["JENIS", "JENIS REKENING", "TIPE"]
TARGET_COLUMN_ALIASES = ["SALDO UPDATE", "SALDO_UPDATE"]
CANONICAL_TARGET_HEADER = "SALDO UPDATE"
# Month-named write targets (SALDO JUNI, SALDO JULI, ...) are a fallback when
# the master already has one. The original document has neither: SALDO UPDATE
# is then appended. The bare historical SALDO column is never a match.
_MONTH_NAMED_SALDO_RE = re.compile(r"^SALDO[_\s]+([A-Z]+)$")
_MONTH_TOKENS = frozenset(
    {
        "JANUARI",
        "FEBRUARI",
        "MARET",
        "APRIL",
        "MEI",
        "JUNI",
        "JULI",
        "AGUSTUS",
        "SEPTEMBER",
        "OKTOBER",
        "NOVEMBER",
        "DESEMBER",
        "JANUARY",
        "FEBRUARY",
        "MARCH",
        "MAY",
        "JUNE",
        "JULY",
        "AUGUST",
        "OCTOBER",
        "DECEMBER",
    }
)
MONTHLY_BALANCE_ALIASES = [
    "SALDO IDR",
    "SALDO_IDR",
    "SALDO IN IDR",
    "SALDO_IN_IDR",
    "SALDO",
]

# JENIS value that is excluded from the update entirely.
DEPOSITO = "deposito"

# Header detection scans every row so a title banner (or a header inserted
# in the middle of the values) still resolves without hardcoded offsets.
_EXCEL_SUFFIXES = (".xlsx", ".xls", ".xlsm")


def _excel_cell_value(value: object) -> object | None:
    """Converts a pandas cell into something openpyxl will store.

    Empty / NaN cells stay blank. Numpy scalars become Python ints/floats so
    account numbers copied out of a legacy .xls still join as digits.
    """
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    item = getattr(value, "item", None)
    if callable(item):
        try:
            return item()
        except (ValueError, AttributeError):
            return value
    return value


REPORT_GIRO_DEFINITION = WorkflowDefinition(
    workflow_id=WorkflowId.REPORT_GIRO,
    label="Report Giro",
    requires_reference=False,
    # The "master" input is the Giro master workbook that gets updated; the
    # "raw" input is the monthly source it is reconciled against.
    requires_master_data=True,
    group_cols=(),
    value_col="SALDO UPDATE",
    report_title="Report Giro",
    detail_sheet_name="",
    # Concatenated digits, no thousands separator and no decimals - the same
    # plain-integer look every other report ships (Portal BG / Vol TF).
    number_format=PLAIN_INTEGER_FORMAT,
    # The monthly source usually carries only NO REK + SALDO IDR, but when it
    # does have SEGMEN/SOURCE/KW the operator gets the same live controls as the
    # other reports. Defaults stay empty so nothing is dropped unasked.
    supports_segment_filter=True,
    has_source_filter=True,
    has_kw_filter=True,
    raw_button_label="1. Monthly Giro File...",
    raw_picker_title="Select the monthly giro source workbook",
    raw_filetypes=(("Excel Workbook", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")),
    master_button_label="2b. Giro Master File...",
    master_picker_title="Select the target giro master workbook",
)


class GiroColumnResolutionError(WorkflowValidationError):
    """Raised when the account/balance columns cannot be resolved by alias."""


class ReportGiroStrategy(WorkflowStrategy):
    """Copies monthly IDR balances into the Giro master's target month column."""

    @property
    def definition(self) -> WorkflowDefinition:
        return REPORT_GIRO_DEFINITION

    def execute(self, config: ProcessingConfig) -> WorkflowRunResult:
        # 0. The master workbook is the thing being updated; without it there is
        # nothing to write into, so fail before reading anything.
        if config.master_data_path is None:
            raise WorkflowValidationError(
                "The 'Report Giro' workflow requires a master-data file (the "
                "Giro master workbook to update), but none was provided."
            )
        master_path = Path(config.master_data_path)
        if not master_path.exists():
            raise WorkflowValidationError(
                f"Giro master workbook not found: {master_path}"
            )

        stages: dict[str, float] = {}

        def mark(name: str, started: float) -> float:
            stages[name] = round(perf_counter() - started, 4)
            return perf_counter()

        # 1. Build the monthly {account -> saldo} lookup (VLOOKUP first hit),
        # honoring the operator's SEGMEN/SOURCE filters where those columns exist.
        # Every sheet is still scanned; aliases still decide the winning columns.
        clock = perf_counter()
        self._emit_progress(config, "Reading monthly source...")
        balances = self._load_monthly_balances(Path(config.raw_data_path), config)
        clock = mark("read", clock)

        # 2. Update the master in memory, then save to the output path. Nothing
        # is written until every row succeeded, so a resolution failure can
        # never leave a partial output that looks like a successful run.
        self._emit_progress(config, "Loading master...")
        workbook = self._load_master_workbook(master_path)
        worksheet, header_row, columns = self._resolve_master_layout(workbook)
        clock = mark("match", clock)

        self._emit_progress(config, "Writing balances...")
        scanned, unmatched = self._apply_balances(
            worksheet, header_row, columns, balances
        )

        output_path = self._xlsx_output_path(Path(config.output_report_path))
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._emit_progress(config, "Saving Excel...")
        try:
            workbook.save(output_path)
        except PermissionError as exc:
            raise ExportError(
                f"Cannot write '{output_path}'. The file may be open in Microsoft "
                f"Excel. Close it or choose a different output path."
            ) from exc
        mark("write", clock)

        return WorkflowRunResult(
            output_path=output_path,
            total_records_processed=scanned,
            unmapped_records_count=unmatched,
            stage_timings=stages,
        )

    # ------------------------------------------------------------------ #
    # Monthly source
    # ------------------------------------------------------------------ #
    def _load_monthly_balances(
        self, monthly_path: Path, config: ProcessingConfig
    ) -> dict[str, float]:
        """Returns {canonical account: saldo} from the monthly source workbook.

        Every sheet is scanned; the first one resolving both an account column
        and a balance column wins. Duplicate accounts keep the first hit,
        matching VLOOKUP.

        The operator's SEGMENT / SOURCE / KW filters are applied to the winning
        sheet *before* the lookup is built, so a filtered account simply never
        matches and its master cell is left as it was. A monthly extract without
        those columns makes the filters a no-op (the Portal BG / Qlola
        precedent), so typing a filter can never fail a run.
        """
        if not monthly_path.exists():
            raise WorkflowValidationError(
                f"Monthly giro source file not found: {monthly_path}"
            )

        frames = self._read_source_frames(monthly_path)
        attempts: list[tuple[str, list[str]]] = []
        for sheet_name, frame in frames:
            account_col = resolve_alias_column(frame, ACCOUNT_ALIASES)
            balance_col = resolve_alias_column(frame, MONTHLY_BALANCE_ALIASES)
            if account_col is None or balance_col is None:
                attempts.append((sheet_name, [str(c) for c in frame.columns]))
                continue
            frame = self.apply_runtime_filters(frame, config)
            return self._mapping_from_columns(frame, account_col, balance_col)

        raise GiroColumnResolutionError(
            "Unable to resolve the account and balance columns in the monthly "
            f"giro source {monthly_path}.\n"
            f"Sheets scanned:\n{self._format_attempts(attempts)}\n"
            "Aliases attempted:\n"
            f"  - account: {ACCOUNT_ALIASES}\n"
            f"  - balance: {MONTHLY_BALANCE_ALIASES}"
        )

    @staticmethod
    def _read_source_frames(path: Path) -> list[tuple[str, pd.DataFrame]]:
        """Reads the monthly source as (sheet name, frame) pairs, all as text.

        Values are read with `dtype=str` so Excel's int/float storage cannot
        mangle the account key before `canonicalize_join_id` sees it.
        """
        suffix = path.suffix.lower()
        try:
            if suffix in _EXCEL_SUFFIXES:
                excel_file = open_excel(path)
                frames = []
                for sheet_name in excel_file.sheet_names:
                    frame = pd.read_excel(
                        excel_file, sheet_name=sheet_name, header=None, dtype=str
                    )
                    frame, _header_row = promote_header_row(
                        frame, (*ACCOUNT_ALIASES, *MONTHLY_BALANCE_ALIASES)
                    )
                    if frame.empty:
                        continue
                    frames.append((str(sheet_name), frame))
                if not frames:
                    raise WorkflowValidationError(EMPTY_INPUT_MESSAGE)
                return frames
            if suffix == ".csv":
                csv_frame = pd.read_csv(path, header=None, dtype=str)
                csv_frame, _header_row = promote_header_row(
                    csv_frame, (*ACCOUNT_ALIASES, *MONTHLY_BALANCE_ALIASES)
                )
                if csv_frame.empty:
                    raise WorkflowValidationError(EMPTY_INPUT_MESSAGE)
                return [("(csv)", csv_frame)]
        except WorkflowValidationError:
            raise
        except Exception as exc:  # noqa: BLE001 - re-wrapped for the caller
            raise WorkflowValidationError(
                f"Failed to read monthly giro source {path}: {exc}"
            ) from exc
        raise WorkflowValidationError(
            f"Unsupported monthly giro source file type: '{suffix}'. "
            "Expected an Excel workbook (.xlsx/.xls/.xlsm) or .csv."
        )

    @staticmethod
    def _mapping_from_columns(
        frame: pd.DataFrame, account_col: str, balance_col: str
    ) -> dict[str, float]:
        """Builds {canonical account: numeric saldo}, first hit wins."""
        pairs = frame[[account_col, balance_col]].copy()
        pairs[account_col] = pairs[account_col].apply(canonicalize_join_id)
        pairs[balance_col] = pd.to_numeric(pairs[balance_col], errors="coerce")
        pairs = pairs[
            (pairs[account_col] != "")
            & (pairs[account_col].str.lower() != "nan")
            & pairs[balance_col].notna()
        ]
        pairs = pairs.drop_duplicates(subset=account_col, keep="first")
        return {
            account: float(saldo)
            for account, saldo in zip(pairs[account_col], pairs[balance_col])
        }

    # ------------------------------------------------------------------ #
    # Master workbook
    # ------------------------------------------------------------------ #
    def _resolve_master_layout(self, workbook) -> tuple[Worksheet, int, dict[str, int]]:
        """Finds the sheet, header row, and 1-based column indexes to work with.

        The first sheet exposing an account column wins. The write target is
        `SALDO UPDATE` when present, else a month-named column such as
        `SALDO JUNI`, else a new `SALDO UPDATE` column appended after the last
        header. Header detection scans the top rows so a title banner above the
        table does not break resolution, and no column letter is hardcoded.
        """
        attempts: list[tuple[str, list[str]]] = []
        for worksheet in workbook.worksheets:
            for header_row in range(1, worksheet.max_row + 1):
                headers = self._header_map(worksheet, header_row)
                if not headers:
                    continue
                account_col = self._alias_index(headers, ACCOUNT_ALIASES)
                if account_col is None:
                    continue
                target_col = self._target_index(headers)
                if target_col is None:
                    target_col = self._append_saldo_update_column(
                        worksheet, header_row, headers
                    )
                return (
                    worksheet,
                    header_row,
                    {
                        "account": account_col,
                        "target": target_col,
                        # JENIS is optional: a master without it simply has no
                        # Deposito rows to skip.
                        "jenis": self._alias_index(headers, JENIS_ALIASES),
                    },
                )
            attempts.append(
                (worksheet.title, list(self._header_map(worksheet, 1).values()))
            )

        raise GiroColumnResolutionError(
            "Unable to resolve the account column in the Giro master workbook.\n"
            f"Sheets scanned:\n{self._format_attempts(attempts)}\n"
            "Aliases attempted:\n"
            f"  - account: {ACCOUNT_ALIASES}\n"
            "The write column is SALDO UPDATE (created when missing) or a "
            "month-named column such as SALDO JUNI / SALDO JULI. Historical "
            "SALDO is never overwritten."
        )

    @staticmethod
    def _header_map(worksheet: Worksheet, row: int) -> dict[int, str]:
        """{1-based column index: header text} for one candidate header row."""
        headers: dict[int, str] = {}
        for column in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row, column).value
            if value is None or str(value).strip() == "":
                continue
            headers[column] = str(value)
        return headers

    @staticmethod
    def _alias_index(headers: dict[int, str], aliases: list[str]) -> int | None:
        """1-based index of the column matching the first alias that is present."""
        by_normalized: dict[str, int] = {}
        for column, text in headers.items():
            by_normalized.setdefault(normalize_header(text), column)
        for alias in aliases:
            match = by_normalized.get(normalize_header(alias))
            if match is not None:
                return match
        return None

    def _target_index(self, headers: dict[int, str]) -> int | None:
        """1-based index of an existing write-target balance column.

        `SALDO UPDATE` / `SALDO_UPDATE` win when present so a master that has
        both the canonical header and a leftover month column is not ambiguous.
        Otherwise the rightmost month-named `SALDO <month>` column is used.
        The historical `SALDO` column is never returned; callers append
        `SALDO UPDATE` when this returns None.
        """
        canonical = self._alias_index(headers, TARGET_COLUMN_ALIASES)
        if canonical is not None:
            return canonical
        month_col = None
        for column, text in headers.items():
            if self._is_month_named_saldo(normalize_header(text)):
                month_col = column
        return month_col

    @staticmethod
    def _append_saldo_update_column(
        worksheet: Worksheet, header_row: int, headers: dict[int, str]
    ) -> int:
        """Creates SALDO UPDATE immediately after the last header cell."""
        target_col = max(headers) + 1
        worksheet.cell(header_row, target_col).value = CANONICAL_TARGET_HEADER
        return target_col

    @staticmethod
    def _is_month_named_saldo(normalized_header: str) -> bool:
        match = _MONTH_NAMED_SALDO_RE.fullmatch(normalized_header)
        return bool(match and match.group(1) in _MONTH_TOKENS)

    def _load_master_workbook(self, path: Path):
        """Loads the Giro master as an openpyxl workbook.

        `.xlsx` / `.xlsm` are opened in place. Legacy `.xls` (the original
        `DAFTAR REKENING GIRO TSPM.xls`) is copied into a new workbook because
        openpyxl cannot edit BIFF8 files.
        """
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xlsm"):
            try:
                return load_workbook(path)
            except Exception as exc:  # noqa: BLE001 - re-wrapped for the caller
                raise WorkflowValidationError(
                    f"Failed to read Giro master workbook {path}: {exc}"
                ) from exc
        if suffix == ".xls":
            return self._workbook_from_legacy_xls(path)
        raise WorkflowValidationError(
            f"Unsupported Giro master file type: '{suffix}'. "
            "Expected an Excel workbook (.xlsx/.xls/.xlsm)."
        )

    def _workbook_from_legacy_xls(self, path: Path) -> Workbook:
        """Copies every sheet of a BIFF8 .xls master into an openpyxl workbook."""
        try:
            excel_file = open_excel(path)
            workbook = Workbook()
            default_sheet = workbook.active
            first = True
            for sheet_name in excel_file.sheet_names:
                frame = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                if first:
                    worksheet = default_sheet
                    worksheet.title = str(sheet_name)[:31]
                    first = False
                else:
                    worksheet = workbook.create_sheet(title=str(sheet_name)[:31])
                self._write_frame_to_worksheet(frame, worksheet)
            return workbook
        except Exception as exc:  # noqa: BLE001 - re-wrapped for the caller
            raise WorkflowValidationError(
                f"Failed to read Giro master workbook {path}: {exc}"
            ) from exc

    @staticmethod
    def _write_frame_to_worksheet(frame: pd.DataFrame, worksheet: Worksheet) -> None:
        for row_idx, row in enumerate(frame.to_numpy(), start=1):
            for col_idx, value in enumerate(row, start=1):
                converted = _excel_cell_value(value)
                if converted is None:
                    continue
                worksheet.cell(row_idx, col_idx).value = converted

    @staticmethod
    def _xlsx_output_path(path: Path) -> Path:
        """openpyxl can only write .xlsx; coerce a leftover .xls output name."""
        if path.suffix.lower() == ".xls":
            return path.with_suffix(".xlsx")
        return path

    def _apply_balances(
        self,
        worksheet: Worksheet,
        header_row: int,
        columns: dict[str, int],
        balances: dict[str, float],
    ) -> tuple[int, int]:
        """Writes matched balances into the target column only.

        Returns (rows scanned, non-Deposito rows with no monthly match). Rows
        that are skipped or unmatched are left exactly as they were - the target
        cell is never blanked, and no sentinel is ever written.
        """
        account_col = columns["account"]
        target_col = columns["target"]
        jenis_col = columns["jenis"]

        scanned = 0
        unmatched = 0
        for row in range(header_row + 1, worksheet.max_row + 1):
            raw_account = worksheet.cell(row, account_col).value
            if raw_account is None or str(raw_account).strip() == "":
                continue  # trailing/blank spacer row, not a record
            scanned += 1

            if jenis_col is not None:
                jenis = worksheet.cell(row, jenis_col).value
                if str(jenis or "").strip().casefold() == DEPOSITO:
                    continue  # Deposito never receives an update

            saldo = balances.get(canonicalize_join_id(raw_account))
            if saldo is None:
                unmatched += 1
                continue  # left blank: no UNMAPPED sentinel in a balance cell

            cell = worksheet.cell(row, target_col)
            # Concatenated digits: rounded to a whole rupiah and written as a
            # real int, so the cell stays numeric rather than becoming a
            # preformatted string.
            cell.value = int(round(saldo))
            # Explicit numeric format so large IDR balances never render as
            # scientific notation or land as General-format text.
            cell.number_format = PLAIN_INTEGER_FORMAT

        return scanned, unmatched

    @staticmethod
    def _format_attempts(attempts: list[tuple[str, list[str]]]) -> str:
        return "\n".join(
            f"  - sheet '{name}': columns {cols}" for name, cols in attempts
        )
