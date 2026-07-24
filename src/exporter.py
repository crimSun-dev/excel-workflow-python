"""Excel Report Exporter (TDD Section 3.5).

Writes a styled two-sheet workbook using openpyxl (no Excel install required):
    * Summary_Report  - tabular pivot view + Grand Total, corporate styling,
                        explicit #,##0.00 number format (no scientific notation).
    * Enriched_Data   - full row-level audit trail.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

VALUE_COLUMN = "VOLUME_IN_IDR"

_DEFAULT_REPORT_TITLE = "Financial Summary Report"
_DEFAULT_DETAIL_SHEET = "Enriched_Data"

# Corporate styling palette (TDD 1.3).
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
_META_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FONT = Font(bold=True)
_THIN_SIDE = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


class ExportError(Exception):
    """Raised when writing or styling the Excel output file fails."""


class ExcelReportExporter:
    """Formats and exports summary and row-level datasets to an openpyxl workbook."""

    def __init__(
        self,
        number_format: str = "#,##0.00",
        value_column: str = VALUE_COLUMN,
        report_title: str = _DEFAULT_REPORT_TITLE,
        detail_sheet_name: str = _DEFAULT_DETAIL_SHEET,
        value_columns: tuple[str, ...] | None = None,
        display_names: dict[str, str] | None = None,
    ):
        self.number_format = number_format
        # Parameterized so each workflow can supply its own value column
        # (VOLUME_IN_IDR vs AMOUNT_IN_IDR), report heading, and detail tab name.
        self.value_column = value_column
        # Ordered value columns to number-format and Grand-Total. Defaults to the
        # single `value_column` so single-value workflows are unaffected.
        self.value_columns = tuple(value_columns) if value_columns else (value_column,)
        # Internal column -> report header overrides (e.g. FBI -> Sum of FBI).
        self.display_names = dict(display_names or {})
        self.report_title = report_title
        self.detail_sheet_name = detail_sheet_name

    def _value_col_indices(self, columns: list[str]) -> set[int]:
        """1-based positions of the value columns present in `columns`."""
        return {
            columns.index(c) + 1 for c in self.value_columns if c in columns
        }

    def export(
        self,
        summary_df: pd.DataFrame,
        enriched_df: pd.DataFrame,
        output_path: Path,
        segment_filter_applied: str | None = None,
    ) -> Path:
        """Writes a workbook with formatted Summary_Report and Enriched_Data tabs.

        Args:
            summary_df: Summarized Pivot Table DataFrame.
            enriched_df: Row-level enriched DataFrame.
            output_path: Destination path for the .xlsx file.
            segment_filter_applied: Filter label for the report header block.

        Returns:
            Path to the written workbook.

        Raises:
            ExportError: If the destination is locked (open in Excel) or
                openpyxl fails to write.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        summary_ws = workbook.active
        summary_ws.title = "Summary_Report"
        self._write_summary_sheet(
            summary_ws, summary_df, segment_filter_applied, len(enriched_df)
        )

        enriched_ws = workbook.create_sheet(self.detail_sheet_name)
        self._write_enriched_sheet(enriched_ws, enriched_df)

        try:
            workbook.save(output_path)
        except PermissionError as exc:
            raise ExportError(
                f"Cannot write '{output_path}'. The file may be open in Microsoft "
                f"Excel. Close it or choose a different output path."
            ) from exc
        except Exception as exc:  # noqa: BLE001 - re-wrapped for the caller
            raise ExportError(f"Failed to write Excel report: {exc}") from exc

        return output_path

    def _write_summary_sheet(
        self,
        ws: Worksheet,
        summary_df: pd.DataFrame,
        segment_filter_applied: str | None,
        total_records: int,
    ) -> None:
        """Writes the metadata block, pivot table, and Grand Total row."""
        # --- Metadata header block ---
        meta = [
            (self.report_title, ""),
            ("Processing Timestamp:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Records Processed:", f"{total_records:,}"),
            ("Filter Applied (SEGMEN):", segment_filter_applied or "None (All Segments)"),
        ]
        for label, value in meta:
            row = ws.max_row + 1 if ws.max_row > 1 or ws["A1"].value else ws.max_row
            ws.cell(row=row, column=1, value=label).font = _TOTAL_FONT
            ws.cell(row=row, column=1).fill = _META_FILL
            ws.cell(row=row, column=2, value=value).fill = _META_FILL
        title_cell = ws["A1"]
        title_cell.font = Font(bold=True, size=14, color="1F4E79")

        header_row = ws.max_row + 2  # blank spacer row before the table
        columns = list(summary_df.columns)

        # --- Column headers (with optional display-name overrides) ---
        for col_idx, col_name in enumerate(columns, start=1):
            header_text = self.display_names.get(col_name, col_name)
            cell = ws.cell(row=header_row, column=col_idx, value=header_text)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

        # --- Data rows ---
        value_col_indices = self._value_col_indices(columns)
        first_data_row = header_row + 1
        for r_offset, (_, record) in enumerate(summary_df.iterrows()):
            excel_row = first_data_row + r_offset
            for col_idx, col_name in enumerate(columns, start=1):
                value = record[col_name]
                cell = ws.cell(row=excel_row, column=col_idx)
                if col_idx in value_col_indices:
                    cell.value = float(value)
                    cell.number_format = self.number_format
                else:
                    cell.value = value
                cell.border = _BORDER

        # --- Grand Total row (totals every value column) ---
        total_row = first_data_row + len(summary_df)
        ws.cell(row=total_row, column=1, value="Grand Total").font = _TOTAL_FONT
        if not summary_df.empty:
            for col_idx in value_col_indices:
                col_name = columns[col_idx - 1]
                total_value = float(summary_df[col_name].sum())
                total_cell = ws.cell(row=total_row, column=col_idx, value=total_value)
                total_cell.number_format = self.number_format
                total_cell.font = _TOTAL_FONT
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill = _TOTAL_FILL
            cell.border = _BORDER

        self._auto_fit_columns(ws)

    def _write_enriched_sheet(self, ws: Worksheet, enriched_df: pd.DataFrame) -> None:
        """Writes the full row-level audit trail with header styling."""
        columns = list(enriched_df.columns)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _BORDER

        value_col_indices = self._value_col_indices(columns)
        for r_offset, (_, record) in enumerate(enriched_df.iterrows()):
            excel_row = 2 + r_offset
            for col_idx, col_name in enumerate(columns, start=1):
                value = record[col_name]
                cell = ws.cell(row=excel_row, column=col_idx)
                if col_idx in value_col_indices:
                    cell.value = float(value) if pd.notna(value) else 0.0
                    cell.number_format = self.number_format
                else:
                    cell.value = value

        self._auto_fit_columns(ws)
        ws.freeze_panes = "A2"

    @staticmethod
    def _auto_fit_columns(ws: Worksheet) -> None:
        """Scales column widths dynamically based on the longest cell string."""
        widths: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                length = len(str(cell.value))
                if cell.number_format and cell.number_format != "General":
                    # Formatted numbers render wider than their raw repr.
                    length = max(length, len(f"{cell.value:,.2f}") + 2)
                widths[cell.column] = max(widths.get(cell.column, 0), length)
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(width + 2, 10), 60
            )
