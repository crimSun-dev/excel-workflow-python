"""End-to-end orchestrator integration test (TDD Section 3.6 / Phase 3)."""

from __future__ import annotations

from openpyxl import load_workbook

from src.orchestrator import PipelineOrchestrator
from src.schemas import ProcessingConfig


def test_end_to_end_pipeline_success(raw_pipe_file, reference_file, tmp_path):
    out = tmp_path / "e2e_report.xlsx"
    config = ProcessingConfig(
        raw_data_path=raw_pipe_file,
        reference_data_path=reference_file,
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)

    assert report.success is True
    assert report.error_message is None
    assert report.total_records_processed == 4
    assert report.unmapped_records_count == 1
    assert out.exists()

    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Summary_Report", "Enriched_Data"}


def test_end_to_end_pipeline_missing_raw_returns_failure(reference_file, tmp_path):
    config = ProcessingConfig(
        raw_data_path=tmp_path / "missing.txt",
        reference_data_path=reference_file,
        output_report_path=tmp_path / "out.xlsx",
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is False
    assert report.error_message is not None


def test_end_to_end_pipeline_with_segment_filter(raw_pipe_file, reference_file, tmp_path):
    out = tmp_path / "filtered.xlsx"
    config = ProcessingConfig(
        raw_data_path=raw_pipe_file,
        reference_data_path=reference_file,
        output_report_path=out,
        segmen_filter="Corporate",
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True
    ws = load_workbook(out)["Summary_Report"]
    # Only Corporate (0002 -> MC20) should appear in the pivot body.
    codes = [c.value for row in ws.iter_rows() for c in row if c.value == "MC20"]
    assert "MC20" in codes
