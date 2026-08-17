"""Multi-workflow tests: strategy dispatch, new workflows, and E2E pipeline."""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.aggregation import AggregationEngine
from src.enrichment import (
    MasterDataEnricher,
    ReferenceEnricher,
    ReferenceEnrichmentError,
    canonicalize_join_id,
    warn_if_mostly_unmapped,
)
from src.gui import (
    default_kw_text,
    default_segmen_exclude_text,
    default_segmen_include_text,
    default_source_text,
    filter_hint_text,
    kw_filter_enabled,
    master_picker_config,
    parse_filter_input,
    raw_picker_config,
    segment_filter_enabled,
    source_filter_enabled,
    unmapped_warning_text,
    workflow_choices,
)
from src.ingestion import IngestionEngine
from src.orchestrator import PipelineOrchestrator
from src.schemas import ProcessingConfig
from src.workflows.base import (
    WorkflowDefinition,
    WorkflowId,
    WorkflowStrategy,
    WorkflowValidationError,
    normalize_join_key,
)
from src.workflows.registry import WORKFLOW_REGISTRY, get_definition, get_strategy


# --------------------------------------------------------------------------- #
# canonicalize_join_id unit tests (Task 1.3)
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("1001", "1001"),
        ("1001.0", "1001"),        # Excel int→float→str round-trip
        ("1001.00", "1001"),       # double trailing zeros
        ("  1001  ", "1001"),      # padded whitespace
        (" 1001.0 ", "1001"),      # padding + trailing .0
        (1001, "1001"),            # actual int input
        (1001.0, "1001"),          # actual float input
        ("U1", "U1"),             # non-numeric preserved
        ("", ""),                  # empty string
        ("1001.5", "1001.5"),      # fractional part preserved
        ("00123", "00123"),        # leading zeros preserved (distinct ID)
    ],
)
def test_canonicalize_join_id(raw, expected):
    assert canonicalize_join_id(raw) == expected


def test_canonicalize_fixes_float_master_id_mismatch(tmp_path):
    """End-to-end: master stores int 1001, raw has string '1001' → match, not UNMAPPED."""
    master = pd.DataFrame({"ID": [1001, 1002], "MAIN_CODE": ["7", "9"]})
    path = tmp_path / "master_int.xlsx"
    master.to_excel(path, index=False)
    mapping = MasterDataEnricher(path).build_id_to_main_code()
    # canonicalize_join_id collapses 1001.0 → "1001"
    assert mapping.get("1001") == "7"
    assert mapping.get("1002") == "9"


# --------------------------------------------------------------------------- #
# Registry / dispatch
# --------------------------------------------------------------------------- #
def test_registry_has_all_seven_workflows():
    assert set(WORKFLOW_REGISTRY) == {
        WorkflowId.AKUMULASI,
        WorkflowId.RINCIAN_VOL_TF,
        WorkflowId.RINCIAN_PORTAL_BG,
        WorkflowId.TIMESERIES_FBI_BRIVA,
        WorkflowId.TIMESERIES_ACTIVE_USER_QLOLA,
        WorkflowId.REPORT_DATA_STATIS,
        WorkflowId.REPORT_GIRO,
    }


def test_registry_covers_every_declared_workflow_id():
    """Every WorkflowId must dispatch, so a new enum member can't be forgotten."""
    assert set(WORKFLOW_REGISTRY) == set(WorkflowId)


# --------------------------------------------------------------------------- #
# Selector order
#
# `WorkflowId` declaration order is the display contract: enums iterate in
# declaration order, so the GUI dropdown, the CLI's valid-options listing, and
# these assertions all read from the same source. The set-based tests above
# would happily pass through a reorder, hence these ordered ones.
# --------------------------------------------------------------------------- #
OPERATOR_SEQUENCE = [
    WorkflowId.AKUMULASI,
    WorkflowId.TIMESERIES_ACTIVE_USER_QLOLA,
    WorkflowId.REPORT_DATA_STATIS,
    WorkflowId.RINCIAN_VOL_TF,
    WorkflowId.RINCIAN_PORTAL_BG,
    WorkflowId.TIMESERIES_FBI_BRIVA,
    WorkflowId.REPORT_GIRO,
]

OPERATOR_SEQUENCE_LABELS = [
    "Report Summary Akumulasi",
    "Time Series Active User Qlola",
    "Report Data Statis",
    "Rincian Vol TF",
    "Rincian Portal BG",
    "Report Vol Briva",
    "Report Giro",
]


def test_workflow_id_declaration_order_matches_operator_sequence():
    assert list(WorkflowId) == OPERATOR_SEQUENCE


def test_gui_dropdown_labels_match_operator_sequence():
    labels = list(workflow_choices())
    assert labels == OPERATOR_SEQUENCE_LABELS
    # The default selection is the first entry; no eighth item may appear.
    assert labels[0] == "Report Summary Akumulasi"
    assert len(labels) == 7


def test_gui_dropdown_labels_map_to_unchanged_ids():
    """Reordering must not rename any id the operator's scripts already use."""
    assert workflow_choices() == {
        "Report Summary Akumulasi": "akumulasi",
        "Time Series Active User Qlola": "timeseries-active-user-qlola",
        "Report Data Statis": "report-data-statis",
        "Rincian Vol TF": "rincian-vol-tf",
        "Rincian Portal BG": "rincian-portal-bg",
        "Report Vol Briva": "timeseries-fbi-briva",
        "Report Giro": "report-giro",
    }


def test_registry_insertion_order_mirrors_the_enum():
    assert list(WORKFLOW_REGISTRY) == OPERATOR_SEQUENCE


def test_cli_valid_options_listing_follows_the_enum_order():
    """The unknown-id error builds its list from WorkflowId, so it stays in sync."""
    from typer.testing import CliRunner

    from src.cli import app

    result = CliRunner().invoke(
        app, ["process", "--raw", "nope.txt", "--workflow", "does-not-exist"]
    )
    assert result.exit_code == 1
    listed = result.output.split("Valid options:")[1]
    positions = [listed.index(w.value) for w in OPERATOR_SEQUENCE]
    assert positions == sorted(positions)


def test_get_strategy_accepts_string_id():
    strategy = get_strategy("rincian-vol-tf")
    assert strategy.definition.workflow_id is WorkflowId.RINCIAN_VOL_TF


def test_get_strategy_resolves_new_timeseries_ids():
    briva = get_strategy("timeseries-fbi-briva")
    assert briva.definition.workflow_id is WorkflowId.TIMESERIES_FBI_BRIVA
    qlola = get_strategy("timeseries-active-user-qlola")
    assert qlola.definition.workflow_id is WorkflowId.TIMESERIES_ACTIVE_USER_QLOLA
    assert qlola.definition.requires_master_data is True


def test_get_strategy_unknown_id_raises():
    with pytest.raises(WorkflowValidationError):
        get_strategy("does-not-exist")


# --------------------------------------------------------------------------- #
# Aggregation exclusion (unit)
# --------------------------------------------------------------------------- #
def test_aggregation_excludes_wholesale_case_insensitive():
    df = pd.DataFrame(
        {
            "SEGMEN": ["wholesale", "Corporate", "", "WHOLESALE"],
            "MAINBR": ["B01", "B01", "B02", "B03"],
            "MBDESC": ["One", "One", "Two", "Three"],
            "AMOUNT_IN_IDR": [1000.0, 2000.0, 300.0, 500.0],
        }
    )
    result = AggregationEngine(exclude_segmen=["Wholesale"]).aggregate(
        df, group_cols=["MAINBR", "MBDESC"], value_col="AMOUNT_IN_IDR"
    )
    # B01 keeps only the Corporate row (2000); B02 keeps the blank-SEGMEN row
    # (300); B03 is dropped entirely (only Wholesale).
    assert result.branch_count == 2
    assert result.total_volume_idr == 2300.0


# --------------------------------------------------------------------------- #
# Count aggregation + regional KW inclusion (unit)
# --------------------------------------------------------------------------- #
def _count_frame() -> pd.DataFrame:
    """MC10 has P1, P1, P2 and a blank ID; MC20 has one ID."""
    return pd.DataFrame(
        {
            "MAIN_CODE": ["MC10", "MC10", "MC10", "MC10", "MC20"],
            "MAIN_BRANCH": ["One", "One", "One", "One", "Two"],
            "ID_PRODUCT": ["P1", "P1", "P2", "   ", "P9"],
        }
    )


def test_count_aggregation_counts_duplicates_and_skips_blanks():
    result = AggregationEngine().aggregate(
        _count_frame(), value_col="ID_PRODUCT", aggfunc="count"
    )
    summary = result.summary_data.set_index("MAIN_CODE")["ID_PRODUCT"]
    # Duplicated P1 counts twice; the whitespace-only cell does not count.
    assert summary["MC10"] == 3
    assert summary["MC20"] == 1
    assert result.total_volume_idr == 4.0
    assert result.branch_count == 2


def test_count_aggregation_yields_integers_not_floats():
    result = AggregationEngine().aggregate(
        _count_frame(), value_col="ID_PRODUCT", aggfunc="count"
    )
    assert pd.api.types.is_integer_dtype(result.summary_data["ID_PRODUCT"])


def test_count_aggregation_ignores_null_ids():
    df = pd.DataFrame(
        {
            "MAIN_CODE": ["MC10", "MC10"],
            "MAIN_BRANCH": ["One", "One"],
            "ID_PRODUCT": ["P1", None],
        }
    )
    result = AggregationEngine().aggregate(
        df, value_col="ID_PRODUCT", aggfunc="count"
    )
    assert result.total_volume_idr == 1.0


def test_count_aggregation_drops_korporasi_and_keeps_blank_segmen():
    df = pd.DataFrame(
        {
            "SEGMEN": ["KORPORASI", "korporasi", "Consumer", "Micro", "SME", ""],
            "MAIN_CODE": ["MC10"] * 6,
            "MAIN_BRANCH": ["One"] * 6,
            "ID_PRODUCT": ["P1", "P2", "P3", "P4", "P5", "P6"],
        }
    )
    result = AggregationEngine(exclude_segmen=["KORPORASI"]).aggregate(
        df, value_col="ID_PRODUCT", aggfunc="count"
    )
    # Both KORPORASI casings drop; Consumer/Micro/SME/blank all count.
    assert result.total_volume_idr == 4.0


def test_regional_kw_include_keeps_only_matching_region_any_case():
    df = pd.DataFrame(
        {
            "KW": ["KANWIL MALANG", " kanwil malang ", "KANWIL SURABAYA", ""],
            "MAIN_CODE": ["MC10"] * 4,
            "MAIN_BRANCH": ["One"] * 4,
            "ID_PRODUCT": ["P1", "P2", "P3", "P4"],
        }
    )
    result = AggregationEngine(kw_include=["KANWIL MALANG"]).aggregate(
        df, value_col="ID_PRODUCT", aggfunc="count"
    )
    # Surabaya and the blank KW row are both excluded.
    assert result.total_volume_idr == 2.0


def test_regional_kw_include_and_exclude_segmen_combine_with_and():
    df = pd.DataFrame(
        {
            "SEGMEN": ["KORPORASI", "Consumer", "Consumer"],
            "KW": ["KANWIL MALANG", "KANWIL MALANG", "KANWIL SURABAYA"],
            "MAIN_CODE": ["MC10"] * 3,
            "MAIN_BRANCH": ["One"] * 3,
            "ID_PRODUCT": ["P1", "P2", "P3"],
        }
    )
    result = AggregationEngine(
        exclude_segmen=["KORPORASI"], kw_include=["KANWIL MALANG"]
    ).aggregate(df, value_col="ID_PRODUCT", aggfunc="count")
    # Only the Consumer + Malang row survives both filters.
    assert result.total_volume_idr == 1.0


def test_regional_kw_include_filtering_everything_out_returns_empty_summary():
    df = pd.DataFrame(
        {
            "KW": ["KANWIL SURABAYA"],
            "MAIN_CODE": ["MC10"],
            "MAIN_BRANCH": ["One"],
            "ID_PRODUCT": ["P1"],
        }
    )
    result = AggregationEngine(kw_include=["KANWIL MALANG"]).aggregate(
        df, value_col="ID_PRODUCT", aggfunc="count"
    )
    assert result.branch_count == 0
    assert result.total_volume_idr == 0.0


def test_sum_remains_the_default_aggfunc():
    df = pd.DataFrame(
        {
            "MAIN_CODE": ["MC10", "MC10"],
            "MAIN_BRANCH": ["One", "One"],
            "VOLUME_IN_IDR": [1000.0, 2000.0],
        }
    )
    result = AggregationEngine().aggregate(df)
    assert result.total_volume_idr == 3000.0


# --------------------------------------------------------------------------- #
# Rincian Vol TF workflow
# --------------------------------------------------------------------------- #
def test_rincian_vol_tf_excludes_wholesale_and_keeps_blank(rincian_vol_tf_file, tmp_path):
    out = tmp_path / "vol_tf.xlsx"
    config = ProcessingConfig(
        raw_data_path=rincian_vol_tf_file,
        workflow_id="rincian-vol-tf",
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message
    assert out.exists()

    ws = load_workbook(out)["Summary_Report"]
    rows = {
        (r[0].value, r[1].value): r[2].value
        for r in ws.iter_rows()
        if r[0].value in {"B01", "B02", "B03"}
    }
    assert rows[("B01", "Branch One")] == 2000  # Wholesale row excluded
    assert rows[("B02", "Branch Two")] == 800   # blank SEGMEN row retained
    assert ("B03", "Branch Three") not in rows  # only-Wholesale branch dropped


def test_rincian_vol_tf_runs_without_reference_file(rincian_vol_tf_file, tmp_path):
    config = ProcessingConfig(
        raw_data_path=rincian_vol_tf_file,
        workflow_id="rincian-vol-tf",
        output_report_path=tmp_path / "no_ref.xlsx",
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message


# --------------------------------------------------------------------------- #
# Rincian Portal BG workflow
# --------------------------------------------------------------------------- #
def test_rincian_portal_bg_aggregates_without_filter(rincian_portal_bg_file, tmp_path):
    out = tmp_path / "portal_bg.xlsx"
    config = ProcessingConfig(
        raw_data_path=rincian_portal_bg_file,
        workflow_id="rincian-portal-bg",
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    rows = {
        (r[0].value, r[1].value): r[2].value
        for r in ws.iter_rows()
        if r[0].value in {"B01", "B02"}
    }
    assert rows[("B01", "Branch One")] == 3000
    assert rows[("B02", "Branch Two")] == 500

    # Plain integer display — no thousands-separator commas (format "0").
    amount_cells = [
        row[2]
        for row in ws.iter_rows()
        if row[0].value in {"B01", "B02"}
    ]
    assert all(isinstance(c.value, int) for c in amount_cells)
    assert all(c.number_format == "0" for c in amount_cells)


# --------------------------------------------------------------------------- #
# Report Data Statis workflow (E2E)
# --------------------------------------------------------------------------- #
def _statis_summary_rows(worksheet) -> dict[tuple, object]:
    """Summary_Report data rows keyed by (MAIN_CODE, MAIN_BRANCH)."""
    return {
        (r[0].value, r[1].value): r[2]
        for r in worksheet.iter_rows()
        if r[0].value in {"MC10", "MC20", "UNMAPPED"}
    }


def _run_data_statis(raw_path: Path, reference: Path, out: Path):
    config = ProcessingConfig(
        raw_data_path=raw_path,
        reference_data_path=reference,
        workflow_id="report-data-statis",
        output_report_path=out,
    )
    return PipelineOrchestrator.execute(config)


def test_report_data_statis_counts_id_product(
    report_data_statis_bin_file, reference_file, tmp_path
):
    out = tmp_path / "data_statis.xlsx"
    report = _run_data_statis(report_data_statis_bin_file, reference_file, out)
    assert report.success is True, report.error_message
    assert out.exists()

    ws = load_workbook(out)["Summary_Report"]
    rows = _statis_summary_rows(ws)
    # MC10: P1, P1, P2 count (duplicates count separately, blank ID does not).
    assert rows[("MC10", "Jakarta Pusat")].value == 3
    # MC20: KANWIL SURABAYA and blank-KW rows dropped, leaving one.
    assert rows[("MC20", "Surabaya")].value == 1
    # 0009 has no reference row but still survives the filters.
    assert rows[("UNMAPPED", "UNMAPPED")].value == 1


def test_report_data_statis_summary_cells_are_plain_integers(
    report_data_statis_bin_file, reference_file, tmp_path
):
    out = tmp_path / "data_statis_format.xlsx"
    _run_data_statis(report_data_statis_bin_file, reference_file, out)

    ws = load_workbook(out)["Summary_Report"]
    cells = list(_statis_summary_rows(ws).values())
    assert cells
    assert all(isinstance(c.value, int) for c in cells)
    assert all(c.number_format == "0" for c in cells)


def test_report_data_statis_header_is_count_of_id_product(
    report_data_statis_bin_file, reference_file, tmp_path
):
    out = tmp_path / "data_statis_header.xlsx"
    _run_data_statis(report_data_statis_bin_file, reference_file, out)

    ws = load_workbook(out)["Summary_Report"]
    headers = [
        [c.value for c in row]
        for row in ws.iter_rows()
        if row[0].value == "MAIN_CODE"
    ]
    assert headers[0][:3] == ["MAIN_CODE", "MAIN_BRANCH", "Count of ID_PRODUCT"]


def test_report_data_statis_exports_two_sheets_with_surviving_rows(
    report_data_statis_bin_file, reference_file, tmp_path
):
    out = tmp_path / "data_statis_sheets.xlsx"
    _run_data_statis(report_data_statis_bin_file, reference_file, out)

    workbook = load_workbook(out)
    assert workbook.sheetnames == ["Summary_Report", "Enriched_Data"]
    detail = workbook["Enriched_Data"]
    # Enriched_Data is the row-level audit trail: every ingested row plus header.
    assert detail.max_row == 11


def test_report_data_statis_bin_parses_identically_to_txt(
    report_data_statis_bin_file, report_data_statis_txt_file, reference_file, tmp_path
):
    bin_out = tmp_path / "from_bin.xlsx"
    txt_out = tmp_path / "from_txt.xlsx"
    _run_data_statis(report_data_statis_bin_file, reference_file, bin_out)
    _run_data_statis(report_data_statis_txt_file, reference_file, txt_out)

    bin_rows = {k: c.value for k, c in _statis_summary_rows(
        load_workbook(bin_out)["Summary_Report"]
    ).items()}
    txt_rows = {k: c.value for k, c in _statis_summary_rows(
        load_workbook(txt_out)["Summary_Report"]
    ).items()}
    assert bin_rows == txt_rows


def test_report_data_statis_requires_reference_file(
    report_data_statis_bin_file, tmp_path
):
    config = ProcessingConfig(
        raw_data_path=report_data_statis_bin_file,
        workflow_id="report-data-statis",
        output_report_path=tmp_path / "no_ref.xlsx",
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is False
    assert "reference" in (report.error_message or "").lower()


def test_report_data_statis_definition_bakes_its_filters():
    definition = get_definition("report-data-statis")
    assert definition.label == "Report Data Statis"
    assert definition.requires_reference is True
    assert definition.aggfunc == "count"
    assert definition.value_col == "ID_PRODUCT"
    assert definition.exclude_segmen == ("KORPORASI",)
    assert definition.number_format == "0"
    # KORPORASI is the default exclusion, but the operator may override it.
    assert definition.supports_segment_filter is True
    # The region is a KW keep-list, not a dimension of its own, and it is an
    # operator control like every other filter.
    assert definition.kw_include == ("KANWIL MALANG",)
    assert definition.has_kw_filter is True
    assert not hasattr(definition, "kawil_include")
    # The region column is the extract's own `KW`, never a `KAWIL` header.
    assert "KW" in definition.required_columns
    assert "KAWIL" not in definition.required_columns
    # ID_PRODUCT must never be coerced to a number by ingestion.
    assert "ID_PRODUCT" not in definition.numeric_columns


def test_report_data_statis_reads_the_region_from_the_kw_column(
    report_data_statis_bin_file, reference_file, tmp_path
):
    """Regression: a real extract spells the region `KW`, and used to fail
    validation demanding an absent `KAWIL` column."""
    out = tmp_path / "data_statis_kw.xlsx"
    report = _run_data_statis(report_data_statis_bin_file, reference_file, out)

    assert report.success is True, report.error_message
    rows = _statis_summary_rows(load_workbook(out)["Summary_Report"])
    # KANWIL SURABAYA and blank-KW rows are still dropped, so the KW keep-list
    # really ran rather than silently passing every region through.
    assert rows[("MC20", "Surabaya")].value == 1


def test_report_data_statis_fails_loudly_without_a_kw_column(
    reference_file, tmp_path
):
    """No region column at all must stop the run, not quietly report every region."""
    raw = tmp_path / "no_kw.bin"
    raw.write_text(
        "KODE_UNIT|SEGMEN|ID_PRODUCT\n0001|Consumer|P1\n", encoding="utf-8"
    )
    report = _run_data_statis(raw, reference_file, tmp_path / "no_kw.xlsx")

    assert report.success is False
    assert "KW" in (report.error_message or "")


def test_report_data_statis_kw_box_overrides_the_default_region(
    report_data_statis_bin_file, reference_file, tmp_path
):
    """A typed KW keep-list replaces KANWIL MALANG for that run."""
    out = tmp_path / "data_statis_surabaya.xlsx"
    config = ProcessingConfig(
        raw_data_path=report_data_statis_bin_file,
        reference_data_path=reference_file,
        workflow_id="report-data-statis",
        output_report_path=out,
        kw_include=["KANWIL SURABAYA"],
    )
    report = PipelineOrchestrator.execute(config)

    assert report.success is True, report.error_message
    rows = _statis_summary_rows(load_workbook(out)["Summary_Report"])
    # Only the single Surabaya row survives; the Malang rows are now filtered out.
    assert {k: c.value for k, c in rows.items()} == {("MC20", "Surabaya"): 1}


def test_report_data_statis_kw_hint_states_the_default_region():
    assert filter_hint_text("report-data-statis", "KW") == "auto: keeps KANWIL MALANG"


def test_report_data_statis_enriches_on_kode_unit_alias(
    report_data_statis_bin_file, reference_file_uker_new, tmp_path
):
    """Raw header `KODE_UNIT` joins a mapping sheet whose key is `KODE UNIT`."""
    out = tmp_path / "data_statis_alias.xlsx"
    report = _run_data_statis(
        report_data_statis_bin_file, reference_file_uker_new, out
    )
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    rows = _statis_summary_rows(ws)
    assert rows[("MC10", "Jakarta Pusat")].value == 3
    assert rows[("MC20", "Surabaya")].value == 1


# --------------------------------------------------------------------------- #
# Report Giro workflow (E2E)
#
# These are the regression guards required by the spec: they must fail if the
# historical SALDO is overwritten, if a Deposito row is filled, or if an
# unmatched Giro account receives anything (including an UNMAPPED sentinel).
# --------------------------------------------------------------------------- #
GIRO_HEADER_ROW = 1
GIRO_SALDO_COL = 3
GIRO_SALDO_UPDATE_COL = 4
# Data rows in the master fixture, in sheet order.
GIRO_ROW_MATCHED = 2      # float-key match  -> 2500
GIRO_ROW_PLAIN = 3        # plain match      -> 7500
GIRO_ROW_DEPOSITO = 4     # in monthly but JENIS=Deposito -> stays blank
GIRO_ROW_MISSING = 5      # absent from monthly            -> stays blank
GIRO_ROW_DUPLICATED = 6   # duplicated in monthly, first hit -> 1250000000000


def _run_giro(monthly: Path, master: Path, out: Path, **filters):
    config = ProcessingConfig(
        raw_data_path=monthly,
        master_data_path=master,
        workflow_id="report-giro",
        output_report_path=out,
        **filters,
    )
    return PipelineOrchestrator.execute(config)


@pytest.fixture
def giro_output(giro_monthly_file, giro_master_file, tmp_path):
    """Runs the workflow once and hands back the updated master worksheet."""
    out = tmp_path / "giro_updated.xlsx"
    report = _run_giro(giro_monthly_file, giro_master_file, out)
    assert report.success is True, report.error_message
    workbook = load_workbook(out)
    return workbook, workbook["DAFTAR GIRO"]


def test_giro_updates_matched_non_deposito_rows(giro_output):
    _, ws = giro_output
    assert ws.cell(GIRO_ROW_MATCHED, GIRO_SALDO_UPDATE_COL).value == 2500
    assert ws.cell(GIRO_ROW_PLAIN, GIRO_SALDO_UPDATE_COL).value == 7500


def test_giro_float_key_variant_still_matches(giro_output):
    """Master key '1234567890' vs monthly float 1234567890.0."""
    _, ws = giro_output
    assert ws.cell(GIRO_ROW_MATCHED, GIRO_SALDO_UPDATE_COL).value == 2500


def test_giro_duplicate_monthly_account_takes_first_hit(giro_output):
    _, ws = giro_output
    assert ws.cell(GIRO_ROW_DUPLICATED, GIRO_SALDO_UPDATE_COL).value == 1250000000000


def test_giro_historical_saldo_column_is_untouched(giro_output):
    """Regression guard: writing to SALDO instead of SALDO UPDATE must fail here."""
    _, ws = giro_output
    saldo = [ws.cell(r, GIRO_SALDO_COL).value for r in range(2, 7)]
    assert saldo == [1000, 500, 300, 900, 100]


def test_giro_deposito_row_is_skipped(giro_output):
    """Regression guard: the Deposito account exists in the monthly file."""
    _, ws = giro_output
    assert ws.cell(GIRO_ROW_DEPOSITO, GIRO_SALDO_UPDATE_COL).value is None
    assert ws.cell(GIRO_ROW_DEPOSITO, GIRO_SALDO_COL).value == 300


def test_giro_unmatched_account_stays_blank(giro_output):
    """Regression guard: no UNMAPPED sentinel may reach a balance cell."""
    _, ws = giro_output
    assert ws.cell(GIRO_ROW_MISSING, GIRO_SALDO_UPDATE_COL).value is None


def test_giro_no_unmapped_sentinel_anywhere_in_target_column(giro_output):
    _, ws = giro_output
    values = [ws.cell(r, GIRO_SALDO_UPDATE_COL).value for r in range(2, 7)]
    assert "UNMAPPED" not in [str(v) for v in values]


def test_giro_updated_cells_are_plain_integers(giro_output):
    """Concatenated digits: no thousands separator, no decimal point, no E+12."""
    _, ws = giro_output
    updated = [
        ws.cell(r, GIRO_SALDO_UPDATE_COL)
        for r in (GIRO_ROW_MATCHED, GIRO_ROW_PLAIN, GIRO_ROW_DUPLICATED)
    ]
    assert all(isinstance(c.value, int) for c in updated)
    assert all(not isinstance(c.value, str) for c in updated)
    assert all(c.number_format == "0" for c in updated)
    # The trillion-scale account stays an exact number, not a float/string.
    assert ws.cell(GIRO_ROW_DUPLICATED, GIRO_SALDO_UPDATE_COL).value == 1250000000000


def test_giro_segmen_exclude_drops_that_account_from_the_lookup(
    giro_monthly_file_with_segmen, giro_master_file, tmp_path
):
    """An excluded segment simply never matches; its master cell is left as it was."""
    out = tmp_path / "giro_segmen_excluded.xlsx"
    report = _run_giro(
        giro_monthly_file_with_segmen,
        giro_master_file,
        out,
        segmen_exclude=["KORPORASI"],
    )
    assert report.success is True, report.error_message
    ws = load_workbook(out)["DAFTAR GIRO"]
    # The KORPORASI account was filtered out of the monthly source.
    assert ws.cell(GIRO_ROW_MATCHED, GIRO_SALDO_UPDATE_COL).value is None
    # Its historical SALDO is still untouched - nothing was blanked.
    assert ws.cell(GIRO_ROW_MATCHED, GIRO_SALDO_COL).value == 1000
    # Kept segments are updated as usual.
    assert ws.cell(GIRO_ROW_PLAIN, GIRO_SALDO_UPDATE_COL).value == 7500


def test_giro_source_exclude_drops_that_account_from_the_lookup(
    giro_monthly_file_with_segmen, giro_master_file, tmp_path
):
    out = tmp_path / "giro_source_excluded.xlsx"
    report = _run_giro(
        giro_monthly_file_with_segmen, giro_master_file, out, source_exclude=["CMS"]
    )
    assert report.success is True, report.error_message
    ws = load_workbook(out)["DAFTAR GIRO"]
    assert ws.cell(GIRO_ROW_DUPLICATED, GIRO_SALDO_UPDATE_COL).value is None
    assert ws.cell(GIRO_ROW_PLAIN, GIRO_SALDO_UPDATE_COL).value == 7500


def test_giro_segmen_keep_only_retains_just_that_segment(
    giro_monthly_file_with_segmen, giro_master_file, tmp_path
):
    out = tmp_path / "giro_segmen_kept.xlsx"
    report = _run_giro(
        giro_monthly_file_with_segmen, giro_master_file, out, segmen_filter="RITEL"
    )
    assert report.success is True, report.error_message
    ws = load_workbook(out)["DAFTAR GIRO"]
    assert ws.cell(GIRO_ROW_MATCHED, GIRO_SALDO_UPDATE_COL).value is None
    assert ws.cell(GIRO_ROW_PLAIN, GIRO_SALDO_UPDATE_COL).value == 7500
    assert ws.cell(GIRO_ROW_DUPLICATED, GIRO_SALDO_UPDATE_COL).value == 1250000000000


def test_giro_filters_are_a_no_op_when_the_monthly_file_lacks_those_columns(
    giro_monthly_file, giro_master_file, tmp_path
):
    """A typical NO REK + SALDO IDR extract cannot match a filter, so nothing drops."""
    out = tmp_path / "giro_filters_noop.xlsx"
    report = _run_giro(
        giro_monthly_file,
        giro_master_file,
        out,
        segmen_filter="RITEL",
        segmen_exclude=["KORPORASI"],
        source_exclude=["CMS"],
    )
    assert report.success is True, report.error_message
    ws = load_workbook(out)["DAFTAR GIRO"]
    assert ws.cell(GIRO_ROW_MATCHED, GIRO_SALDO_UPDATE_COL).value == 2500
    assert ws.cell(GIRO_ROW_PLAIN, GIRO_SALDO_UPDATE_COL).value == 7500
    assert ws.cell(GIRO_ROW_DUPLICATED, GIRO_SALDO_UPDATE_COL).value == 1250000000000


def test_giro_preserves_other_sheets_and_headers(giro_output):
    workbook, ws = giro_output
    assert workbook.sheetnames == ["DAFTAR GIRO", "Catatan"]
    assert workbook["Catatan"]["A1"].value == "Catatan operator"
    assert workbook["Catatan"]["A2"].value == "Jangan dihapus"
    headers = [ws.cell(GIRO_HEADER_ROW, c).value for c in range(1, 5)]
    # The target contract is the month-neutral SALDO UPDATE, never SALDO JUNI.
    assert headers == ["NO REK", "JENIS", "SALDO", "SALDO UPDATE"]


def test_giro_month_named_target_column_is_not_the_contract(
    giro_monthly_file, tmp_path
):
    """A master offering only SALDO JUNI fails loudly rather than being written."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DAFTAR GIRO"
    sheet.append(["NO REK", "JENIS", "SALDO", "SALDO JUNI"])
    sheet.append(["1234567890", "Giro", 1000, None])
    master = tmp_path / "giro_master_juni.xlsx"
    workbook.save(master)

    out = tmp_path / "giro_juni.xlsx"
    report = _run_giro(giro_monthly_file, master, out)
    assert report.success is False
    assert "SALDO UPDATE" in report.error_message
    assert "Sheets scanned" in report.error_message
    assert out.exists() is False


def test_giro_does_not_produce_a_summary_report_sheet(giro_output):
    workbook, _ = giro_output
    assert "Summary_Report" not in workbook.sheetnames


def test_giro_leaves_the_source_master_file_unmodified(
    giro_monthly_file, giro_master_file, tmp_path
):
    """The master is the template; the run writes to the output path only."""
    before = giro_master_file.read_bytes()
    _run_giro(giro_monthly_file, giro_master_file, tmp_path / "giro_copy.xlsx")
    assert giro_master_file.read_bytes() == before


def test_giro_reports_unmatched_account_count(
    giro_monthly_file, giro_master_file, tmp_path
):
    report = _run_giro(
        giro_monthly_file, giro_master_file, tmp_path / "giro_telemetry.xlsx"
    )
    assert report.success is True, report.error_message
    # 5 master data rows scanned; only the missing Giro account is unmatched
    # (the Deposito row is skipped by rule, not by a failed lookup).
    assert report.total_records_processed == 5
    assert report.unmapped_records_count == 1


def test_giro_requires_master_file(giro_monthly_file, tmp_path):
    config = ProcessingConfig(
        raw_data_path=giro_monthly_file,
        workflow_id="report-giro",
        output_report_path=tmp_path / "giro_no_master.xlsx",
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is False
    assert "master" in (report.error_message or "").lower()


def test_giro_unresolvable_monthly_columns_fail_loudly(
    giro_monthly_file_unresolvable, giro_master_file, tmp_path
):
    out = tmp_path / "giro_unresolvable.xlsx"
    report = _run_giro(giro_monthly_file_unresolvable, giro_master_file, out)
    assert report.success is False
    message = report.error_message or ""
    # The error must name the sheets scanned and the aliases attempted.
    assert "Sheet" in message
    assert "NO REK" in message
    # A failed run must not leave a partial output that looks successful.
    assert not out.exists()


def test_giro_gui_offers_excel_pickers_not_pipe_delimited_labels():
    button, title, filetypes = raw_picker_config("report-giro")
    assert "Giro" in button
    assert "pipe-delimited" not in title.lower()
    patterns = " ".join(pattern for _, pattern in filetypes)
    assert "*.xlsx" in patterns and "*.xls" in patterns

    master_button, master_title = master_picker_config("report-giro")
    assert "Giro" in master_button
    # Qlola's "ID -> MAIN_CODE" wording would misdescribe the Giro master.
    assert "MAIN_CODE" not in master_title


def test_other_workflows_keep_the_text_raw_picker():
    _, title, filetypes = raw_picker_config("akumulasi")
    assert "pipe-delimited" in title.lower()
    assert "*.txt" in " ".join(pattern for _, pattern in filetypes)


def test_giro_exposes_segmen_and_source_filters_like_the_others():
    """Giro applies them to its monthly source, so the fields must be live."""
    assert segment_filter_enabled("report-giro") is True
    assert source_filter_enabled("report-giro") is True
    assert kw_filter_enabled("report-giro") is True


def test_every_filter_dimension_is_enabled_for_every_workflow():
    """The unified FILTERS block is live on every report, without exception."""
    for workflow_id in WorkflowId:
        assert segment_filter_enabled(workflow_id) is True
        assert source_filter_enabled(workflow_id) is True
        assert kw_filter_enabled(workflow_id) is True


def test_giro_cli_rejects_a_missing_master_path(giro_monthly_file, tmp_path):
    from typer.testing import CliRunner

    from src.cli import app

    result = CliRunner().invoke(
        app,
        [
            "process",
            "--raw", str(giro_monthly_file),
            "--workflow", "report-giro",
            "--out", str(tmp_path / "giro_cli.xlsx"),
        ],
    )
    assert result.exit_code == 1
    assert "--master" in result.output
    assert "giro master" in result.output.lower()


def test_giro_cli_runs_with_both_paths(giro_monthly_file, giro_master_file, tmp_path):
    from typer.testing import CliRunner

    from src.cli import app

    out = tmp_path / "giro_cli_ok.xlsx"
    result = CliRunner().invoke(
        app,
        [
            "process",
            "--raw", str(giro_monthly_file),
            "--master", str(giro_master_file),
            "--workflow", "report-giro",
            "--out", str(out),
        ],
    )
    assert result.exit_code == 0, result.output
    assert out.exists()


def test_giro_definition_declares_its_inputs():
    definition = get_definition("report-giro")
    assert definition.label == "Report Giro"
    assert definition.requires_master_data is True
    assert definition.requires_reference is False
    # Plain integers, matching every other report - not "#,##0.00".
    assert definition.number_format == "0"
    # The filters are live with empty defaults; nothing is dropped unasked.
    assert definition.supports_segment_filter is True
    assert definition.has_source_filter is True
    assert definition.segmen_include is None
    assert definition.exclude_segmen == ()
    assert definition.source_exclude == ()


# --------------------------------------------------------------------------- #
# Validation & E2E structure
# --------------------------------------------------------------------------- #
def test_missing_required_column_reports_failure(tmp_path):
    # Portal BG file missing the MBNAME column.
    bad = tmp_path / "bad.csv"
    bad.write_text("MAINBR|AMOUNT_IN_IDR\nB01|1000\n", encoding="utf-8")
    config = ProcessingConfig(
        raw_data_path=bad,
        workflow_id="rincian-portal-bg",
        output_report_path=tmp_path / "out.xlsx",
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is False
    assert "MBNAME" in (report.error_message or "")


def test_new_workflow_uses_detail_sheet_name(rincian_portal_bg_file, tmp_path):
    out = tmp_path / "sheets.xlsx"
    config = ProcessingConfig(
        raw_data_path=rincian_portal_bg_file,
        workflow_id="rincian-portal-bg",
        output_report_path=out,
    )
    PipelineOrchestrator.execute(config)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Summary_Report", "Detail_Data"}


def test_akumulasi_workflow_unchanged_end_to_end(raw_pipe_file, reference_file, tmp_path):
    """Regression: default workflow still produces the original two sheets."""
    out = tmp_path / "akum.xlsx"
    config = ProcessingConfig(
        raw_data_path=raw_pipe_file,
        reference_data_path=reference_file,
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True
    assert report.total_records_processed == 4
    assert report.unmapped_records_count == 1
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Summary_Report", "Enriched_Data"}


# --------------------------------------------------------------------------- #
# Akumulasi Sum of FBI
# --------------------------------------------------------------------------- #
def _read_summary_table(ws):
    """Extracts (headers, {MAIN_CODE: row-dict}, grand-total-dict) from a sheet."""
    header_row = None
    for row in ws.iter_rows():
        if row[0].value == "MAIN_CODE":
            header_row = row[0].row
            break
    assert header_row is not None, "Summary table header not found"

    headers = []
    for cell in ws[header_row]:
        if cell.value is None:
            break
        headers.append(cell.value)

    rows: dict[str, dict] = {}
    grand_total: dict = {}
    r = header_row + 1
    while True:
        first = ws.cell(row=r, column=1).value
        if first is None:
            break
        record = {
            headers[i]: ws.cell(row=r, column=i + 1).value for i in range(len(headers))
        }
        if first == "Grand Total":
            grand_total = record
            break
        rows[first] = record
        r += 1
    return headers, rows, grand_total


def test_akumulasi_summary_has_sum_of_fbi_left_of_volume(
    raw_pipe_file, reference_file, tmp_path
):
    out = tmp_path / "akum_fbi.xlsx"
    config = ProcessingConfig(
        raw_data_path=raw_pipe_file,
        reference_data_path=reference_file,
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    headers, _, _ = _read_summary_table(ws)
    assert "Sum of FBI" in headers
    assert headers.index("Sum of FBI") == headers.index("VOLUME_IN_IDR") - 1


def test_akumulasi_sums_fbi_and_volume_with_grand_totals(
    raw_pipe_file, reference_file, tmp_path
):
    out = tmp_path / "akum_totals.xlsx"
    config = ProcessingConfig(
        raw_data_path=raw_pipe_file,
        reference_data_path=reference_file,
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    _, rows, grand_total = _read_summary_table(ws)

    # FBI branch sums export as plain integers (standard round).
    assert rows["MC10"]["Sum of FBI"] == 300
    assert rows["MC10"]["VOLUME_IN_IDR"] == 1_500_000_000_000
    assert rows["MC20"]["Sum of FBI"] == 50
    assert rows["MC20"]["VOLUME_IN_IDR"] == 500_000_000
    assert rows["UNMAPPED"]["Sum of FBI"] == 10

    # Both Grand Totals present as plain ints.
    assert grand_total["Sum of FBI"] == 360
    assert grand_total["VOLUME_IN_IDR"] == int(
        round(1_500_501_000_000.50)
    )


def test_akumulasi_missing_fbi_fails_with_clear_error(tmp_path):
    bad = tmp_path / "no_fbi.txt"
    bad.write_text(
        "KODE_UKER|SEGMEN|VOLUME_IN_IDR\n0001|Wholesale|1000\n", encoding="utf-8"
    )
    config = ProcessingConfig(
        raw_data_path=bad,
        workflow_id="akumulasi",
        output_report_path=tmp_path / "out.xlsx",
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is False
    assert "FBI" in (report.error_message or "")
    assert not (tmp_path / "out.xlsx").exists()


def test_ingestion_coerces_amount_column():
    engine = IngestionEngine(numeric_columns=("AMOUNT_IN_IDR",))
    # Round-trip a small file via a DataFrame-style check on coercion.
    df = pd.DataFrame({"AMOUNT_IN_IDR": ["1,000", "2000", "bad"]})
    coerced = engine._coerce_numeric_columns(df.copy())
    assert coerced["AMOUNT_IN_IDR"].tolist() == [1000.0, 2000.0, 0.0]


# --------------------------------------------------------------------------- #
# Reference enrichment extensions (raw alias + master ID->MAIN_CODE)
# --------------------------------------------------------------------------- #
def test_enrichment_resolves_raw_kode_uker_alias(reference_file):
    # Raw frame carries the spaced `KODE UKER` while the lookup key is KODE_UKER.
    raw = pd.DataFrame({"KODE UKER": ["0001", "0002"], "VOLUME_IDR": [1.0, 2.0]})
    result = ReferenceEnricher(reference_file).enrich(raw)
    assert "MAIN_CODE" in result.data.columns
    assert "MAIN_BRANCH" in result.data.columns
    assert result.unmapped_count == 0


def test_master_enricher_maps_id_to_main_code(qlola_master_file):
    mapping = MasterDataEnricher(qlola_master_file).build_id_to_main_code()
    assert mapping["U1"] == "7"
    assert mapping["U3"] == "9"
    assert "U5" not in mapping  # absent from master => later marked UNMAPPED


def test_master_enricher_unresolvable_columns_raises(tmp_path):
    bad = pd.DataFrame({"FOO": ["1"], "BAR": ["2"]})
    path = tmp_path / "master_bad.xlsx"
    bad.to_excel(path, index=False)
    with pytest.raises(ReferenceEnrichmentError) as exc:
        MasterDataEnricher(path).build_id_to_main_code()
    message = str(exc.value)
    assert "Aliases attempted" in message
    assert "Sheets scanned" in message


def test_master_enricher_master_data_statis_headers(tmp_path):
    """Task 2.1: fixture mimicking MASTER DATA STATIS MALANG workbook headers.

    Real file has columns like ID, MAIN_CODE (or KODE INDUK) — the alias
    registry must resolve them. This verifies the current ID_ALIASES and
    MAIN_CODE_ALIASES cover the master workbook without needing index-21 API.
    """
    master = pd.DataFrame({
        "ID": ["1001", "1002", "1003"],
        "MAIN_CODE": ["7", "9", "7"],
    })
    path = tmp_path / "master_statis_malang.xlsx"
    master.to_excel(path, sheet_name="MASTER DATA STATIS MALANG", index=False)
    mapping = MasterDataEnricher(path).build_id_to_main_code()
    assert mapping["1001"] == "7"
    assert mapping["1002"] == "9"
    assert mapping["1003"] == "7"


def test_master_enricher_ambiguous_two_sheet_picks_correct_sheet(tmp_path):
    """Task 2.3: ambiguous two-sheet master → correct sheet wins.

    Sheet A resolves aliases but maps everything to WRONG codes.
    Sheet B (named "MASTER DATA") is the real mapping table.
    Preference scoring must pick sheet B over sheet A.
    """
    wrong_sheet = pd.DataFrame({
        "ID": ["1001", "1002"],
        "KODE": ["WRONG_A", "WRONG_B"],  # KODE resolves as MAIN_CODE alias
    })
    correct_sheet = pd.DataFrame({
        "ID": ["1001", "1002"],
        "MAIN_CODE": ["7", "9"],
    })
    path = tmp_path / "master_ambiguous.xlsx"
    with pd.ExcelWriter(path) as writer:
        wrong_sheet.to_excel(writer, sheet_name="Cover", index=False)
        correct_sheet.to_excel(writer, sheet_name="MASTER DATA", index=False)
    mapping = MasterDataEnricher(path).build_id_to_main_code()
    # Must pick the "MASTER DATA" sheet, not the "Cover" sheet.
    assert mapping["1001"] == "7"
    assert mapping["1002"] == "9"


def test_master_enricher_picks_id_product_over_cif(tmp_path):
    """Live MASTER DATA STATIS layout: Qlola IDs in ID_PRODUCT, not CIF."""
    master = pd.DataFrame(
        {
            "ID_PRODUCT": ["100362", "100540", "100583"],
            "CIF": ["R555823", "TK34387", "Z064299"],
            "MAIN CODE": ["9", "429", "44"],
        }
    )
    path = tmp_path / "master_statis_real_layout.xlsx"
    master.to_excel(path, sheet_name="Sheet1", index=False)
    mapping = MasterDataEnricher(path).build_id_to_main_code(
        probe_ids=["100362", "100540"]
    )
    assert mapping["100362"] == "9"
    assert mapping["100540"] == "429"
    assert "R555823" not in mapping


def test_master_enricher_picks_id_column_matching_probe_ids(tmp_path):
    """Decoy ``ID`` (CIF/account codes) must lose to the real user-ID column.

    Mirrors live ``MASTER DATA STATIS MALANG`` workbooks where a generic ``ID``
    column holds values like R555823 while Qlola user IDs live in ``ID QLOLA``.
    """
    master = pd.DataFrame(
        {
            "ID": ["R555823", "TK34387", "Z064299", "M261301"],
            "ID QLOLA": ["100362", "100540", "100583", "100586"],
            "MAIN_CODE": ["73", "110", "9", "33"],
        }
    )
    path = tmp_path / "master_decoy_id.xlsx"
    master.to_excel(path, sheet_name="MASTER DATA STATIS MALANG", index=False)
    mapping = MasterDataEnricher(path).build_id_to_main_code(
        probe_ids=["100362", "100540", "100999"]
    )
    assert mapping["100362"] == "73"
    assert mapping["100540"] == "110"
    assert "R555823" not in mapping


def test_qlola_decoy_master_id_column_maps_via_probe_ids(
    qlola_raw_file_numeric_ids, qlola_uker_reference_file, tmp_path
):
    """End-to-end: master with decoy ID column still produces a real Summary."""
    master = pd.DataFrame(
        {
            "ID": ["R555823", "TK34387", "Z064299"],
            "ID QLOLA": ["1001", "1002", "1003"],
            "MAIN_CODE": ["7", "9", "7"],
        }
    )
    master_path = tmp_path / "master_decoy.xlsx"
    master.to_excel(master_path, sheet_name="MASTER DATA STATIS MALANG", index=False)
    out = tmp_path / "qlola_decoy.xlsx"
    report = _run_qlola(
        qlola_raw_file_numeric_ids, qlola_uker_reference_file, master_path, out
    )
    assert report.success is True, report.error_message
    assert report.unmapped_records_count == 0
    assert report.unmapped_diagnostic is None

    _, rows, _ = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    assert rows["7"]["AKTIF TRX >=5x"] == 2  # 1001 + 1003
    assert rows["9"]["TIDAK AKTIF <5x"] == 1  # 1002


def test_master_enricher_integration_live_like_master_qlola(
    tmp_path, qlola_uker_reference_file
):
    """Task 2.4: live-like master + Qlola sample → Summary has real MAIN_CODEs.

    Simulates a realistic scenario where the master has int-stored IDs
    (triggering the .0 float mismatch) and the raw file has string IDs.
    After canonicalization and preference scoring, the Summary must contain
    real MAIN_CODE rows, not only UNMAPPED.
    """
    # Master with int IDs (will become "1001.0" strings in Excel)
    master_df = pd.DataFrame({
        "ID": [1001, 1002, 1003],
        "MAIN_CODE": ["7", "9", "7"],
    })
    master_path = tmp_path / "master_live_like.xlsx"
    master_df.to_excel(master_path, sheet_name="MASTER DATA STATIS", index=False)

    # Qlola raw with string IDs
    raw_path = tmp_path / "qlola_live.csv"
    raw_path.write_text(
        "SOURCE|ID|FREKUENSI|KODE_UKER\n"
        "QCASH|1001|6|K1\n"
        "QCASH|1002|2|K1\n"
        "QCASH|1003|7|K2\n",
        encoding="utf-8",
    )
    out = tmp_path / "qlola_live.xlsx"
    report = _run_qlola(raw_path, qlola_uker_reference_file, master_path, out)
    assert report.success is True, report.error_message

    _, rows, _ = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    # Must have real MAIN_CODE rows, not only UNMAPPED.
    assert "7" in rows
    assert "9" in rows


# --------------------------------------------------------------------------- #
# Report Vol Briva (timeseries-fbi-briva)
# --------------------------------------------------------------------------- #
def test_briva_nonwholesale_volume_matches_sample_grand_total(
    timeseries_briva_file, reference_file, tmp_path
):
    out = tmp_path / "briva.xlsx"
    config = ProcessingConfig(
        raw_data_path=timeseries_briva_file,
        reference_data_path=reference_file,
        workflow_id="timeseries-fbi-briva",
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    headers, rows, grand_total = _read_summary_table(ws)

    # Value header renders as the sample Sheet1 label.
    assert "Sum of VOLUME_IDR" in headers
    # NONWHOLESALE branch totals (the WHOLESALE row is excluded), ceiled to ints.
    assert rows["MC10"]["Sum of VOLUME_IDR"] == 1_481_517_421_604
    assert rows["MC20"]["Sum of VOLUME_IDR"] == 1_000_000_000_000
    # Grand Total matches Excel ROUND/CEIL of the sample oracle constant.
    assert grand_total["Sum of VOLUME_IDR"] == 2_481_517_421_604


def _grand_total_cell(ws, header: str):
    """Returns the Grand Total row's cell under `header` on a summary sheet."""
    header_row = next(r[0].row for r in ws.iter_rows() if r[0].value == "MAIN_CODE")
    col = next(
        c.column for c in ws[header_row] if c.value == header
    )
    r = header_row + 1
    while ws.cell(row=r, column=1).value is not None:
        if ws.cell(row=r, column=1).value == "Grand Total":
            return ws.cell(row=r, column=col)
        r += 1
    raise AssertionError("Grand Total row not found")


def test_briva_grand_total_rounds_fractional_cents_up(
    timeseries_briva_file, reference_file, tmp_path
):
    """WF4 exports plain integers with fractional cents rounded up (ceil).

    The raw oracle sum is ``…603.75``; manual Excel ``ROUND(..., 0)`` and our
    ``ceil`` both land on ``…604`` — stored as a plain int with format ``0``.
    """
    out = tmp_path / "briva_precision.xlsx"
    config = ProcessingConfig(
        raw_data_path=timeseries_briva_file,
        reference_data_path=reference_file,
        workflow_id="timeseries-fbi-briva",
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    cell = _grand_total_cell(
        load_workbook(out)["Summary_Report"], "Sum of VOLUME_IDR"
    )
    assert cell.value == 2_481_517_421_604
    assert isinstance(cell.value, int)
    assert cell.number_format == "0"


def test_briva_missing_reference_fails_clearly(timeseries_briva_file, tmp_path):
    config = ProcessingConfig(
        raw_data_path=timeseries_briva_file,
        workflow_id="timeseries-fbi-briva",
        output_report_path=tmp_path / "out.xlsx",
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is False
    assert "reference" in (report.error_message or "").lower()
    assert not (tmp_path / "out.xlsx").exists()


# --------------------------------------------------------------------------- #
# Time Series Active User Qlola
# --------------------------------------------------------------------------- #
def _read_crosstab_table(ws):
    """Extracts ({MAIN_CODE: {col: count}}, grand-total-dict) from a crosstab sheet."""
    header_row = None
    for row in ws.iter_rows():
        if row[0].value == "MAIN_CODE":
            header_row = row[0].row
            break
    assert header_row is not None, "Crosstab header not found"

    headers = []
    for cell in ws[header_row]:
        if cell.value is None:
            break
        headers.append(cell.value)

    rows: dict[str, dict] = {}
    grand_total: dict = {}
    r = header_row + 1
    while True:
        first = ws.cell(row=r, column=1).value
        if first is None:
            break
        record = {
            headers[i]: ws.cell(row=r, column=i + 1).value for i in range(len(headers))
        }
        if first == "Grand Total":
            grand_total = record
            break
        rows[first] = record
        r += 1
    return headers, rows, grand_total


def _run_qlola(qlola_raw_file, uker, master, out):
    config = ProcessingConfig(
        raw_data_path=qlola_raw_file,
        reference_data_path=uker,
        master_data_path=master,
        workflow_id="timeseries-active-user-qlola",
        output_report_path=out,
    )
    return PipelineOrchestrator.execute(config)


def test_qlola_crosstab_uses_master_main_code_not_uker(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    out = tmp_path / "qlola.xlsx"
    report = _run_qlola(qlola_raw_file, qlola_uker_reference_file, qlola_master_file, out)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    headers, rows, grand_total = _read_crosstab_table(ws)

    assert headers == ["MAIN_CODE", "AKTIF TRX >=5x", "TIDAK AKTIF <5x", "Grand Total"]

    # Rows are keyed on MASTER MAIN_CODE (7 / 9), NOT the UKER code 99.
    # UNMAPPED is excluded from Summary rows (mapped-only). Detail retains it.
    assert set(rows) == {"7", "9"}
    assert "99" not in rows
    assert "UNMAPPED" not in rows

    # Master-keyed distribution (CMS excluded, FREKUENSI summed per ID, >=5 active):
    #   7: U1(7)=AKTIF, U2(2)+U4(1)=TIDAK  -> 1 / 2
    #   9: U3(5)=AKTIF                      -> 1 / 0  (U6 was CMS-only => dropped)
    #   (UNMAPPED: U5(9)=AKTIF — excluded from Summary)
    assert rows["7"]["AKTIF TRX >=5x"] == 1
    assert rows["7"]["TIDAK AKTIF <5x"] == 2
    assert rows["9"]["AKTIF TRX >=5x"] == 1
    assert rows["9"]["TIDAK AKTIF <5x"] == 0

    # Detail keeps UKER MAIN_CODE (reference) separate from master MAIN_CODE.
    detail_ws = load_workbook(out)["Enriched_Data"]
    headers = [detail_ws.cell(row=1, column=c).value for c in range(1, detail_ws.max_column + 1)]
    assert headers[:5] == [
        "ID",
        "UKER_MAIN_CODE",
        "MAIN_CODE",
        "FREKUENSI",
        "USER_AKTIF",
    ]
    uker_col = headers.index("UKER_MAIN_CODE") + 1
    master_col = headers.index("MAIN_CODE") + 1
    for r in range(2, detail_ws.max_row + 1):
        assert detail_ws.cell(row=r, column=uker_col).value == "99"
        master_code = detail_ws.cell(row=r, column=master_col).value
        assert master_code in {"7", "9", "UNMAPPED"}

    # Grand Totals exclude UNMAPPED (mapped-only: 4, not 5).
    assert grand_total["AKTIF TRX >=5x"] == 2
    assert grand_total["TIDAK AKTIF <5x"] == 2
    assert grand_total["Grand Total"] == 4


def test_qlola_missing_master_fails_clearly(
    qlola_raw_file, qlola_uker_reference_file, tmp_path
):
    out = tmp_path / "qlola_no_master.xlsx"
    config = ProcessingConfig(
        raw_data_path=qlola_raw_file,
        reference_data_path=qlola_uker_reference_file,
        workflow_id="timeseries-active-user-qlola",
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is False
    assert "master" in (report.error_message or "").lower()
    assert not out.exists()


def test_qlola_missing_reference_fails_clearly(
    qlola_raw_file, qlola_master_file, tmp_path
):
    out = tmp_path / "qlola_no_ref.xlsx"
    config = ProcessingConfig(
        raw_data_path=qlola_raw_file,
        master_data_path=qlola_master_file,
        workflow_id="timeseries-active-user-qlola",
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is False
    assert "reference" in (report.error_message or "").lower()
    assert not out.exists()


# --------------------------------------------------------------------------- #
# Reference-side UKER alias resolution
# --------------------------------------------------------------------------- #
@pytest.fixture
def _uker_raw_frame() -> pd.DataFrame:
    return pd.DataFrame({"KODE_UKER": ["0001", "0002"], "VOLUME_IDR": [1.0, 2.0]})


def test_reference_uker_alias_resolves(reference_file_uker, _uker_raw_frame):
    result = ReferenceEnricher(reference_file_uker).enrich(_uker_raw_frame)
    assert result.unmapped_count == 0
    assert result.data["MAIN_CODE"].tolist() == ["MC10", "MC20"]
    assert result.data["MAIN_BRANCH"].tolist() == ["Jakarta Pusat", "Surabaya"]


def test_reference_spaced_kode_uker_alias_resolves(
    reference_file_kode_uker_spaced, _uker_raw_frame
):
    result = ReferenceEnricher(reference_file_kode_uker_spaced).enrich(_uker_raw_frame)
    assert result.unmapped_count == 0
    assert result.data["MAIN_CODE"].tolist() == ["MC10", "MC20"]


def test_reference_canonical_kode_uker_still_resolves(reference_file, _uker_raw_frame):
    """Regression: the canonical header keeps working after the alias widening."""
    result = ReferenceEnricher(reference_file).enrich(_uker_raw_frame)
    assert result.unmapped_count == 0
    assert result.data["MAIN_CODE"].tolist() == ["MC10", "MC20"]


def test_reference_unresolvable_error_lists_columns_and_aliases(
    reference_file_unresolvable, _uker_raw_frame
):
    with pytest.raises(ReferenceEnrichmentError) as exc:
        ReferenceEnricher(reference_file_unresolvable).enrich(_uker_raw_frame)
    message = str(exc.value)
    assert "Sheets scanned" in message
    assert "ACCOUNT NUMBER" in message  # available columns of the scanned sheet
    assert "Aliases attempted" in message
    for alias in ("KODE_UKER", "KODE UKER", "UKER", "KANCA"):
        assert alias in message


# --------------------------------------------------------------------------- #
# Raw-side join key normalisation (workflow ingest step)
# --------------------------------------------------------------------------- #
def test_normalize_join_key_renames_spaced_alias():
    raw = pd.DataFrame({"KODE UKER": ["0001"], "VOLUME_IDR": [1.0]})
    out = normalize_join_key(raw, "KODE_UKER")
    assert "KODE_UKER" in out.columns
    assert "KODE UKER" not in out.columns


def test_normalize_join_key_prefers_kode_uker_over_bare_uker():
    """Real Briva exports carry both; `UKER` holds branch *names*, not codes.

    Regression for the false-UNMAPPED collapse: resolving to `UKER` matched
    'KC Tulungagung' against numeric reference keys and mapped nothing.
    """
    raw = pd.DataFrame(
        {
            "KODE UKER": ["110", "13"],
            "UKER": ["KC Tulungagung", "KC Bondowoso"],
            "VOLUME_IDR": [1.0, 2.0],
        }
    )
    out = normalize_join_key(raw, "KODE_UKER")
    assert out["KODE_UKER"].tolist() == ["110", "13"]
    assert out["UKER"].tolist() == ["KC Tulungagung", "KC Bondowoso"]


def test_enricher_prefers_kode_uker_over_bare_uker_on_both_sides(tmp_path):
    """The same code-beats-name priority applies to the reference workbook."""
    ref = pd.DataFrame(
        {
            "KODE UKER": ["110", "13"],
            "UKER": ["KC Tulungagung", "KC Bondowoso"],
            "MAIN_CODE": ["110", "13"],
            "MAIN_BRANCH": ["KC Tulungagung", "KC Bondowoso"],
        }
    )
    ref_path = tmp_path / "reference_code_and_name.xlsx"
    ref.to_excel(ref_path, index=False)

    raw = pd.DataFrame(
        {"KODE UKER": ["110", "13"], "UKER": ["KC Tulungagung", "KC Bondowoso"]}
    )
    result = ReferenceEnricher(ref_path).enrich(normalize_join_key(raw, "KODE_UKER"))
    assert result.unmapped_count == 0
    assert result.data["MAIN_CODE"].tolist() == ["110", "13"]


def test_normalize_join_key_casts_integer_codes_to_string():
    raw = pd.DataFrame({"UKER": [1, 2], "VOLUME_IDR": [1.0, 2.0]})
    out = normalize_join_key(raw, "KODE_UKER")
    assert out["KODE_UKER"].tolist() == ["1", "2"]


def test_normalize_join_key_strips_padding_and_leaves_source_untouched():
    raw = pd.DataFrame({"KODE_UKER": ["  0001 ", "0002"]})
    out = normalize_join_key(raw, "KODE_UKER")
    assert out["KODE_UKER"].tolist() == ["0001", "0002"]
    # The caller's frame is never mutated in place.
    assert raw["KODE_UKER"].tolist() == ["  0001 ", "0002"]


def test_normalize_join_key_without_any_alias_is_a_noop():
    raw = pd.DataFrame({"BRANCH_ID": ["0001"]})
    out = normalize_join_key(raw, "KODE_UKER")
    assert list(out.columns) == ["BRANCH_ID"]


def test_briva_integer_branch_codes_still_map(reference_file):
    """Numeric-typed raw codes join against the string reference keys."""
    raw = pd.DataFrame({"KODE UKER": [1, 2], "VOLUME_IDR": [1.0, 2.0]})
    normalized = normalize_join_key(raw, "KODE_UKER")
    ref = pd.DataFrame(
        {
            "KODE_UKER": ["1", "2"],
            "MAIN_CODE": ["MC10", "MC20"],
            "MAIN_BRANCH": ["A", "B"],
        }
    )
    ref_path = reference_file.parent / "reference_numeric.xlsx"
    ref.to_excel(ref_path, index=False)
    result = ReferenceEnricher(ref_path).enrich(normalized)
    assert result.unmapped_count == 0


# --------------------------------------------------------------------------- #
# WF4 end-to-end mapping
# --------------------------------------------------------------------------- #
def _run_briva(raw, reference, out):
    config = ProcessingConfig(
        raw_data_path=raw,
        reference_data_path=reference,
        workflow_id="timeseries-fbi-briva",
        output_report_path=out,
    )
    return PipelineOrchestrator.execute(config)


def test_briva_underscore_header_matches_spaced_header_totals(
    timeseries_briva_file_underscore_key, reference_file, tmp_path
):
    out = tmp_path / "briva_underscore.xlsx"
    report = _run_briva(timeseries_briva_file_underscore_key, reference_file, out)
    assert report.success is True, report.error_message

    _, rows, grand_total = _read_summary_table(load_workbook(out)["Summary_Report"])
    assert rows["MC10"]["Sum of VOLUME_IDR"] == 1_481_517_421_604
    assert rows["MC20"]["Sum of VOLUME_IDR"] == 1_000_000_000_000
    assert grand_total["Sum of VOLUME_IDR"] == 2_481_517_421_604


def test_briva_grand_total_equals_sum_of_data_rows(
    timeseries_briva_file, reference_file, tmp_path
):
    out = tmp_path / "briva_total.xlsx"
    report = _run_briva(timeseries_briva_file, reference_file, out)
    assert report.success is True, report.error_message

    _, rows, grand_total = _read_summary_table(load_workbook(out)["Summary_Report"])
    row_sum = sum(r["Sum of VOLUME_IDR"] for r in rows.values())
    assert grand_total["Sum of VOLUME_IDR"] == pytest.approx(row_sum)


def test_briva_unmapped_rows_are_retained_not_dropped(
    timeseries_briva_file_with_unmapped, reference_file, tmp_path
):
    out = tmp_path / "briva_unmapped.xlsx"
    report = _run_briva(timeseries_briva_file_with_unmapped, reference_file, out)
    assert report.success is True, report.error_message

    _, rows, grand_total = _read_summary_table(load_workbook(out)["Summary_Report"])
    # Mapped branches still resolve...
    assert rows["MC10"]["Sum of VOLUME_IDR"] == 100
    assert rows["MC20"]["Sum of VOLUME_IDR"] == 200
    # ...and the genuinely unknown branch code survives as UNMAPPED.
    assert rows["UNMAPPED"]["Sum of VOLUME_IDR"] == 50
    assert grand_total["Sum of VOLUME_IDR"] == 350
    assert report.unmapped_records_count == 1


# --------------------------------------------------------------------------- #
# WF5 ID -> MAIN_CODE lookup normalisation
# --------------------------------------------------------------------------- #
def test_qlola_numeric_master_ids_map_correctly(
    qlola_raw_file_numeric_ids,
    qlola_uker_reference_file,
    qlola_master_numeric_file,
    tmp_path,
):
    """Int-dtype master IDs vs. string raw IDs must not collapse to UNMAPPED."""
    out = tmp_path / "qlola_numeric.xlsx"
    report = _run_qlola(
        qlola_raw_file_numeric_ids,
        qlola_uker_reference_file,
        qlola_master_numeric_file,
        out,
    )
    assert report.success is True, report.error_message

    _, rows, _ = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    assert rows["7"]["AKTIF TRX >=5x"] == 1       # 1001, FREKUENSI 6
    assert rows["9"]["TIDAK AKTIF <5x"] == 1      # 1002, FREKUENSI 2
    # 1003 absent from master -> UNMAPPED, excluded from Summary but on detail.
    assert "UNMAPPED" not in rows
    # Detail sheet still has UNMAPPED IDs.
    detail_ws = load_workbook(out)["Enriched_Data"]
    detail_ids = [detail_ws.cell(row=r, column=1).value for r in range(2, detail_ws.max_row + 1)]
    assert "1003" in detail_ids


def test_master_enricher_strips_whitespace_padded_ids(tmp_path):
    padded = tmp_path / "master_padded.csv"
    padded.write_text("ID,MAIN_CODE\n  U1  ,  7 \nU2,9\n", encoding="utf-8")
    mapping = MasterDataEnricher(padded).build_id_to_main_code()
    assert mapping["U1"] == "7"
    assert mapping["U2"] == "9"


def test_qlola_padding_only_id_duplicates_count_as_one_user(
    tmp_path, qlola_uker_reference_file, qlola_master_file
):
    """End-to-end invariant: ' U1' and 'U1' are one user whose FREKUENSI sums.

    Ingestion already trims fields, and the strategy re-normalises IDs before
    the per-ID groupby; this pins the resulting behaviour either way.
    """
    raw = tmp_path / "qlola_padded_ids.csv"
    raw.write_text(
        "SOURCE|ID|FREKUENSI|KODE_UKER\nQCASH| U1 |3|K1\nQCASH|U1|4|K1\n",
        encoding="utf-8",
    )
    out = tmp_path / "qlola_padded.xlsx"
    report = _run_qlola(raw, qlola_uker_reference_file, qlola_master_file, out)
    assert report.success is True, report.error_message

    _, rows, grand_total = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    # One user totalling 7 => a single AKTIF row, not two TIDAK AKTIF rows.
    assert rows["7"]["AKTIF TRX >=5x"] == 1
    assert grand_total["Grand Total"] == 1


def test_qlola_missing_master_id_becomes_unmapped(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    out = tmp_path / "qlola_unmapped.xlsx"
    report = _run_qlola(qlola_raw_file, qlola_uker_reference_file, qlola_master_file, out)
    assert report.success is True, report.error_message

    _, rows, _ = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    # U5 is absent from the master mapping -> UNMAPPED, excluded from Summary.
    assert "UNMAPPED" not in rows
    assert report.unmapped_records_count == 1

    # Detail sheet retains UNMAPPED IDs for audit.
    detail_ws = load_workbook(out)["Enriched_Data"]
    headers = [detail_ws.cell(row=1, column=c).value for c in range(1, detail_ws.max_column + 1)]
    main_code_col = headers.index("MAIN_CODE") + 1
    detail_codes = [
        detail_ws.cell(row=r, column=main_code_col).value
        for r in range(2, detail_ws.max_row + 1)
    ]
    assert "UNMAPPED" in detail_codes


# --------------------------------------------------------------------------- #
# WF5 mapped-only Summary (Tasks 3.3 / 3.4)
# --------------------------------------------------------------------------- #
def _read_metadata(ws) -> dict[str, object]:
    """Extracts the {label: value} metadata block above a Summary crosstab."""
    meta: dict[str, object] = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label == "MAIN_CODE":  # crosstab header reached
            break
        if isinstance(label, str) and label.endswith(":"):
            meta[label] = ws.cell(row=r, column=2).value
    return meta


def test_qlola_summary_metadata_reports_unmapped_id_count(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    """Task 3.3: unmapped volume is surfaced in metadata, not as a Summary row."""
    out = tmp_path / "qlola_meta.xlsx"
    report = _run_qlola(qlola_raw_file, qlola_uker_reference_file, qlola_master_file, out)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    meta = _read_metadata(ws)
    # U5 is the only ID absent from the master mapping.
    assert meta["Unmapped IDs:"] == "1"


def test_qlola_summary_metadata_omits_unmapped_line_when_all_mapped(
    qlola_uker_reference_file, qlola_master_file, tmp_path
):
    """No `Unmapped IDs:` line when every ID resolves — avoids a noisy zero row."""
    raw = tmp_path / "qlola_all_mapped.csv"
    raw.write_text(
        "SOURCE|ID|FREKUENSI|KODE_UKER\nQCASH|U1|6|K1\nQCASH|U2|2|K2\n",
        encoding="utf-8",
    )
    out = tmp_path / "qlola_all_mapped.xlsx"
    report = _run_qlola(raw, qlola_uker_reference_file, qlola_master_file, out)
    assert report.success is True, report.error_message

    meta = _read_metadata(load_workbook(out)["Summary_Report"])
    assert "Unmapped IDs:" not in meta


def test_qlola_mapped_only_aktif_total_excludes_unmapped_at_scale(
    qlola_uker_reference_file, tmp_path
):
    """Task 3.4: Clairine's dashboard number — 3,243 AKTIF − 204 unmapped = 3,039.

    Reproduces the live shape at scale: every ID is AKTIF (FREKUENSI >= 5), but
    204 of them are absent from the master mapping. The Summary Grand Total must
    report the mapped-only 3,039, while the detail sheet keeps all 3,243 IDs and
    the metadata reports the 204 that were dropped.
    """
    total_aktif = 3243
    unmapped_aktif = 204
    mapped_aktif = total_aktif - unmapped_aktif
    assert mapped_aktif == 3039

    ids = [f"{100000 + i}" for i in range(total_aktif)]
    mapped_ids = ids[:mapped_aktif]

    raw = tmp_path / "qlola_scale.csv"
    raw.write_text(
        "SOURCE|ID|FREKUENSI|KODE_UKER\n"
        + "".join(f"QCASH|{i}|7|K1\n" for i in ids),
        encoding="utf-8",
    )
    # Two master branches so the Summary has real multi-row structure.
    master = tmp_path / "qlola_scale_master.xlsx"
    pd.DataFrame(
        {
            "ID": mapped_ids,
            "MAIN_CODE": ["7" if n % 2 == 0 else "9" for n in range(mapped_aktif)],
        }
    ).to_excel(master, index=False)

    out = tmp_path / "qlola_scale.xlsx"
    report = _run_qlola(raw, qlola_uker_reference_file, master, out)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    _, rows, grand_total = _read_crosstab_table(ws)

    # Mapped-only rows and totals: the 204 unmapped IDs never reach the Summary.
    assert "UNMAPPED" not in rows
    assert set(rows) == {"7", "9"}
    assert grand_total["AKTIF TRX >=5x"] == 3039
    assert grand_total["TIDAK AKTIF <5x"] == 0
    assert grand_total["Grand Total"] == 3039

    # Metadata and the run result still account for the dropped IDs.
    assert _read_metadata(ws)["Unmapped IDs:"] == f"{unmapped_aktif:,}"
    assert report.unmapped_records_count == unmapped_aktif

    # Detail sheet retains the full 3,243-ID population for audit.
    detail_ws = load_workbook(out)["Enriched_Data"]
    assert detail_ws.max_row - 1 == total_aktif


# --------------------------------------------------------------------------- #
# WF5 SOURCE filter config
# --------------------------------------------------------------------------- #
def _run_qlola_with_source(raw, uker, master, out, source_exclude):
    config = ProcessingConfig(
        raw_data_path=raw,
        reference_data_path=uker,
        master_data_path=master,
        workflow_id="timeseries-active-user-qlola",
        output_report_path=out,
        source_exclude=source_exclude,
    )
    return PipelineOrchestrator.execute(config)


def test_qlola_definition_declares_source_filter_defaults():
    definition = get_definition(WorkflowId.TIMESERIES_ACTIVE_USER_QLOLA)
    assert definition.has_source_filter is True
    assert definition.source_exclude == ("CMS",)


def test_qlola_default_config_excludes_cms(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    out = tmp_path / "qlola_default_source.xlsx"
    report = _run_qlola(qlola_raw_file, qlola_uker_reference_file, qlola_master_file, out)
    assert report.success is True, report.error_message

    _, rows, grand_total = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    # U6 is CMS-only, so it never reaches any MAIN_CODE group.
    # U5 is UNMAPPED (excluded from Summary). Mapped-only total = 4.
    assert grand_total["Grand Total"] == 4
    assert rows["9"]["AKTIF TRX >=5x"] == 1  # U3 only; U6 dropped


def test_qlola_empty_source_exclude_keeps_every_row(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    out = tmp_path / "qlola_no_source_filter.xlsx"
    report = _run_qlola_with_source(
        qlola_raw_file, qlola_uker_reference_file, qlola_master_file, out, []
    )
    assert report.success is True, report.error_message

    _, rows, grand_total = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    # U6 (CMS-only, FREKUENSI 10) now counts under master code 9.
    assert rows["9"]["AKTIF TRX >=5x"] == 2
    # U5 is UNMAPPED (excluded from Summary). Mapped-only total = 5.
    assert grand_total["Grand Total"] == 5


def test_qlola_source_exclude_match_is_case_insensitive(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    out = tmp_path / "qlola_lowercase_source.xlsx"
    report = _run_qlola_with_source(
        qlola_raw_file, qlola_uker_reference_file, qlola_master_file, out, ["cms"]
    )
    assert report.success is True, report.error_message

    _, _, grand_total = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    assert grand_total["Grand Total"] == 4  # identical to the uppercase default (mapped-only)


def test_source_exclude_is_a_no_op_without_a_source_column(
    rincian_portal_bg_file, tmp_path
):
    """A SOURCE value must not alter an extract that has no SOURCE column."""
    out = tmp_path / "portal_bg_with_source.xlsx"
    config = ProcessingConfig(
        raw_data_path=rincian_portal_bg_file,
        workflow_id="rincian-portal-bg",
        output_report_path=out,
        source_exclude=["CMS"],
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    rows = {
        (r[0].value, r[1].value): r[2].value
        for r in ws.iter_rows()
        if r[0].value in {"B01", "B02"}
    }
    assert rows[("B01", "Branch One")] == 3000
    assert rows[("B02", "Branch Two")] == 500


def test_only_qlola_drops_a_source_value_by_default():
    """Every workflow offers the field; only Qlola pre-fills it with a drop."""
    for workflow_id in (
        WorkflowId.AKUMULASI,
        WorkflowId.RINCIAN_VOL_TF,
        WorkflowId.RINCIAN_PORTAL_BG,
        WorkflowId.TIMESERIES_FBI_BRIVA,
    ):
        definition = get_definition(workflow_id)
        assert definition.has_source_filter is True
        assert definition.source_exclude == ()


# --------------------------------------------------------------------------- #
# GUI filter fields / config propagation
# --------------------------------------------------------------------------- #
def test_gui_source_field_default_text():
    assert default_source_text(WorkflowId.TIMESERIES_ACTIVE_USER_QLOLA) == "CMS"
    assert default_source_text(WorkflowId.AKUMULASI) == ""


def test_gui_workflow_defaults_back_the_filter_hints():
    """The boxes ship empty, so the hint is what keeps the baked rule visible."""
    assert default_segmen_exclude_text("report-data-statis") == "KORPORASI"
    assert default_segmen_exclude_text("rincian-vol-tf") == "Wholesale"
    assert default_segmen_include_text("timeseries-fbi-briva") == "NONWHOLESALE"
    # No default rule, but the field is still live.
    assert default_segmen_include_text("akumulasi") == ""
    assert default_segmen_exclude_text("akumulasi") == ""
    # Report Giro's fields are live - it just drops nothing by default.
    assert default_segmen_exclude_text("report-giro") == ""
    assert default_segmen_include_text("report-giro") == ""
    # Report Data Statis is the only workflow that bakes in a KW keep-list: its
    # regional cut. Everywhere else the dimension is operator-only.
    assert default_kw_text("report-data-statis") == "KANWIL MALANG"
    for workflow_id in WorkflowId:
        if workflow_id is not WorkflowId.REPORT_DATA_STATIS:
            assert default_kw_text(workflow_id) == ""


@pytest.mark.parametrize(
    ("workflow_id", "dimension", "expected"),
    [
        ("report-data-statis", "SEGMENT", "auto: drops KORPORASI"),
        ("timeseries-fbi-briva", "SEGMENT", "auto: keeps NONWHOLESALE"),
        ("timeseries-active-user-qlola", "SOURCE", "auto: drops CMS"),
        ("akumulasi", "SEGMENT", "auto: keeps everything"),
        ("report-giro", "KW", "auto: keeps everything"),
    ],
)
def test_gui_filter_hint_states_what_an_empty_box_does(
    workflow_id, dimension, expected
):
    assert filter_hint_text(workflow_id, dimension) == expected


def test_gui_filter_hint_marks_a_dimension_the_report_ignores(monkeypatch):
    """A greyed-out box must say why, not just sit there blank."""
    import src.gui as gui

    monkeypatch.setattr(gui, "get_definition", lambda _: _FLAGLESS_DEFINITION)
    for dimension in ("SEGMENT", "SOURCE", "KW"):
        assert gui.filter_hint_text("report-giro", dimension) == (
            "not used by this report"
        )


def test_gui_filter_hint_rejects_an_unknown_dimension():
    with pytest.raises(ValueError):
        filter_hint_text("akumulasi", "KAWIL")


@pytest.mark.parametrize(
    ("text", "expected"),
    [
        ("CMS", ["CMS"]),
        ("CMS,MOBILE", ["CMS", "MOBILE"]),
        ("  CMS , MOBILE  ", ["CMS", "MOBILE"]),
        ("CMS,,MOBILE", ["CMS", "MOBILE"]),
        # An empty box is "run the report's automatic rules", not "filter nothing",
        # so it must reach the workflow as None rather than an empty list.
        ("", None),
        ("   ", None),
        (",  ,", None),
    ],
)
def test_gui_parse_filter_input(text, expected):
    assert parse_filter_input(text) == expected


def test_gui_edited_source_field_reaches_the_workflow(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    """A typed SOURCE box becomes a keep-list and drops the baked CMS exclusion."""
    strategy = get_strategy(WorkflowId.TIMESERIES_ACTIVE_USER_QLOLA)
    typed = parse_filter_input("CMS, MOBILE")
    config = ProcessingConfig(
        raw_data_path=qlola_raw_file,
        reference_data_path=qlola_uker_reference_file,
        master_data_path=qlola_master_file,
        workflow_id="timeseries-active-user-qlola",
        output_report_path=tmp_path / "unused.xlsx",
        source_include=typed,
        source_exclude=None if typed is None else [],
    )
    assert strategy.resolve_source_include(config) == ["CMS", "MOBILE"]
    assert strategy.resolve_source_exclude(config) == []


_FLAGLESS_DEFINITION = WorkflowDefinition(
    workflow_id=WorkflowId.REPORT_GIRO,
    label="Flagless",
    requires_reference=False,
    group_cols=(),
    value_col="VALUE",
    report_title="Flagless",
    detail_sheet_name="",
)


class _FlaglessStrategy(WorkflowStrategy):
    """Declares no filter flag at all, exercising the base-class opt-in guard.

    Every shipped workflow now exposes all three filters, so the guard that stops
    a stale GUI/CLI value from leaking into a workflow that never opted in is
    covered by a stub rather than pinned to whichever report happens to opt out.
    """

    @property
    def definition(self) -> WorkflowDefinition:
        return _FLAGLESS_DEFINITION


def _flagless_config(tmp_path, **overrides) -> ProcessingConfig:
    return ProcessingConfig(
        raw_data_path=tmp_path / "raw.xlsx",
        master_data_path=tmp_path / "master.xlsx",
        workflow_id="report-giro",
        output_report_path=tmp_path / "out.xlsx",
        **overrides,
    )


def test_workflow_without_flag_resolves_empty_source_exclude(tmp_path):
    strategy = _FlaglessStrategy()
    config = _flagless_config(tmp_path, source_exclude=["CMS"])
    assert strategy.resolve_source_exclude(config) == []


def test_giro_resolves_the_operator_source_exclude(tmp_path):
    """Giro opted in, so its SOURCE field is honored rather than discarded."""
    strategy = get_strategy(WorkflowId.REPORT_GIRO)
    config = _flagless_config(tmp_path, source_exclude=["CMS", " MOBILE "])
    assert strategy.resolve_source_exclude(config) == ["CMS", "MOBILE"]


# --------------------------------------------------------------------------- #
# Runtime SEGMEN overrides (operator choice beats the workflow default)
# --------------------------------------------------------------------------- #
def _statis_config(raw, reference, out, **overrides):
    return ProcessingConfig(
        raw_data_path=raw,
        reference_data_path=reference,
        workflow_id="report-data-statis",
        output_report_path=out,
        **overrides,
    )


def test_operator_can_keep_korporasi_by_clearing_the_exclusion(
    report_data_statis_bin_file, reference_file, tmp_path
):
    """An empty exclusion list means 'drop nothing', re-including KORPORASI."""
    out = tmp_path / "statis_with_korporasi.xlsx"
    report = PipelineOrchestrator.execute(
        _statis_config(
            report_data_statis_bin_file, reference_file, out, segmen_exclude=[]
        )
    )
    assert report.success is True, report.error_message

    rows = _statis_summary_rows(load_workbook(out)["Summary_Report"])
    # The two KORPORASI rows (P4, P5) now join the default count of 3.
    assert rows[("MC10", "Jakarta Pusat")].value == 5


def test_operator_can_replace_the_default_exclusion(
    report_data_statis_bin_file, reference_file, tmp_path
):
    out = tmp_path / "statis_drop_consumer.xlsx"
    report = PipelineOrchestrator.execute(
        _statis_config(
            report_data_statis_bin_file,
            reference_file,
            out,
            segmen_exclude=["Consumer"],
        )
    )
    assert report.success is True, report.error_message

    rows = _statis_summary_rows(load_workbook(out)["Summary_Report"])
    # KORPORASI (P4, P5) is kept and the Consumer row (P1) is dropped instead.
    assert rows[("MC10", "Jakarta Pusat")].value == 4


def test_operator_can_keep_only_one_segment(
    report_data_statis_bin_file, reference_file, tmp_path
):
    out = tmp_path / "statis_consumer_only.xlsx"
    report = PipelineOrchestrator.execute(
        _statis_config(
            report_data_statis_bin_file,
            reference_file,
            out,
            segmen_filter="consumer",
            segmen_exclude=[],
        )
    )
    assert report.success is True, report.error_message

    rows = _statis_summary_rows(load_workbook(out)["Summary_Report"])
    # Only the Consumer + KANWIL MALANG row survives under MC10.
    assert rows[("MC10", "Jakarta Pusat")].value == 1


def test_untouched_filters_keep_the_workflow_defaults(
    report_data_statis_bin_file, reference_file, tmp_path
):
    """None (nothing supplied) must still drop KORPORASI, as before."""
    out = tmp_path / "statis_defaults.xlsx"
    report = PipelineOrchestrator.execute(
        _statis_config(report_data_statis_bin_file, reference_file, out)
    )
    assert report.success is True, report.error_message

    rows = _statis_summary_rows(load_workbook(out)["Summary_Report"])
    assert rows[("MC10", "Jakarta Pusat")].value == 3


def test_briva_operator_can_clear_the_nonwholesale_inclusion(tmp_path):
    strategy = get_strategy(WorkflowId.TIMESERIES_FBI_BRIVA)
    base = {
        "raw_data_path": tmp_path / "raw.csv",
        "workflow_id": "timeseries-fbi-briva",
        "output_report_path": tmp_path / "out.xlsx",
    }
    # Untouched: the definition default still applies.
    assert strategy.resolve_segment_filter(ProcessingConfig(**base)) == "NONWHOLESALE"
    # Cleared field: every segment contributes.
    assert (
        strategy.resolve_segment_filter(ProcessingConfig(**base, segmen_filter=""))
        is None
    )
    # Edited field: the operator's value wins over the baked default.
    assert (
        strategy.resolve_segment_filter(
            ProcessingConfig(**base, segmen_filter="WHOLESALE")
        )
        == "WHOLESALE"
    )


def test_segmen_overrides_ignored_where_the_fields_are_disabled(tmp_path):
    strategy = _FlaglessStrategy()
    config = _flagless_config(
        tmp_path, segmen_filter="Consumer", segmen_exclude=["KORPORASI"]
    )
    assert strategy.resolve_segment_filter(config) is None
    assert strategy.resolve_segmen_exclude(config) == []


def test_giro_resolves_the_operator_segmen_overrides(tmp_path):
    """Giro opted in, so its SEGMEN fields reach the monthly-source filter."""
    strategy = get_strategy(WorkflowId.REPORT_GIRO)
    config = _flagless_config(
        tmp_path, segmen_filter="Consumer", segmen_exclude=["KORPORASI"]
    )
    assert strategy.resolve_segment_filter(config) == "Consumer"
    assert strategy.resolve_segmen_exclude(config) == ["KORPORASI"]


# --------------------------------------------------------------------------- #
# UNMAPPED diagnostic warning
# --------------------------------------------------------------------------- #
def _warn_args(**overrides):
    args = {
        "unmapped_count": 10,
        "total": 10,
        "raw_key_sample": ["0001", "0002"],
        "reference_key_sample": ["A1", "A2"],
        "aliases_attempted": ["KODE_UKER", "KODE UKER"],
        "context": "Reference enrichment on 'KODE_UKER'",
    }
    args.update(overrides)
    return args


def test_full_unmapped_join_emits_diagnostic_warning(caplog):
    with caplog.at_level(logging.WARNING, logger="src.enrichment"):
        warned = warn_if_mostly_unmapped(**_warn_args())
    assert warned is True
    message = caplog.text
    assert "0001" in message            # raw key sample
    assert "A1" in message              # reference key sample
    assert "KODE UKER" in message       # aliases attempted
    assert "UNMAPPED" in message


def test_half_unmapped_join_does_not_warn(caplog):
    with caplog.at_level(logging.WARNING, logger="src.enrichment"):
        warned = warn_if_mostly_unmapped(**_warn_args(unmapped_count=5, total=10))
    assert warned is False
    assert caplog.text == ""


def test_empty_frame_does_not_warn():
    assert warn_if_mostly_unmapped(**_warn_args(unmapped_count=0, total=0)) is False


def test_unmapped_warning_does_not_abort_the_pipeline(
    timeseries_briva_file, tmp_path, caplog
):
    """A reference that matches nothing still produces a full report."""
    ref = pd.DataFrame(
        {
            "KODE_UKER": ["ZZZZ"],
            "MAIN_CODE": ["MC99"],
            "MAIN_BRANCH": ["Nowhere"],
        }
    )
    ref_path = tmp_path / "reference_no_overlap.xlsx"
    ref.to_excel(ref_path, index=False)

    out = tmp_path / "briva_all_unmapped.xlsx"
    with caplog.at_level(logging.WARNING, logger="src.enrichment"):
        report = _run_briva(timeseries_briva_file, ref_path, out)

    assert report.success is True, report.error_message
    assert out.exists()
    assert "resolved to UNMAPPED" in caplog.text

    _, rows, _ = _read_summary_table(load_workbook(out)["Summary_Report"])
    assert set(rows) == {"UNMAPPED"}


# --------------------------------------------------------------------------- #
# Operator-visible near-total UNMAPPED diagnostic (Tasks 4.1 / 4.2)
# --------------------------------------------------------------------------- #
@pytest.fixture
def qlola_master_no_overlap_file(tmp_path):
    """Master workbook whose IDs match none of the raw Qlola IDs."""
    path = tmp_path / "qlola_master_no_overlap.xlsx"
    pd.DataFrame({"ID": ["ZZ1", "ZZ2"], "MAIN_CODE": ["7", "9"]}).to_excel(
        path, index=False
    )
    return path


def test_qlola_total_unmapped_join_surfaces_operator_diagnostic(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_no_overlap_file, tmp_path
):
    """Task 4.1: a 100% UNMAPPED master join reaches the operator, not just the log."""
    out = tmp_path / "qlola_no_overlap.xlsx"
    report = _run_qlola(
        qlola_raw_file, qlola_uker_reference_file, qlola_master_no_overlap_file, out
    )
    assert report.success is True, report.error_message

    diagnostic = report.unmapped_diagnostic
    assert diagnostic is not None
    assert "U1" in diagnostic                       # raw ID sample
    assert "ZZ1" in diagnostic                      # master ID sample
    assert "ID QLOLA" in diagnostic                 # aliases attempted
    assert "UNMAPPED" in diagnostic

    # The GUI turns that into an actionable warning box.
    message = unmapped_warning_text(report)
    assert message is not None
    assert "master-data file" in message
    assert diagnostic in message


def test_qlola_total_unmapped_join_still_exports(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_no_overlap_file, tmp_path
):
    """Task 4.2 / PROGRESS-003: orphan keys never abort the export."""
    out = tmp_path / "qlola_no_overlap_export.xlsx"
    report = _run_qlola(
        qlola_raw_file, qlola_uker_reference_file, qlola_master_no_overlap_file, out
    )
    assert report.success is True, report.error_message
    assert out.exists()

    workbook = load_workbook(out)
    # Summary is mapped-only, so a fully-unmapped run has no data rows...
    _, rows, _ = _read_crosstab_table(workbook["Summary_Report"])
    assert rows == {}
    # ...but every ID is still on the detail sheet for audit.
    detail_ws = workbook["Enriched_Data"]
    assert detail_ws.max_row - 1 == 5  # U1..U5 post-CMS-filter


def test_qlola_detail_numbers_are_plain_integers(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    """Task 6.1/6.2: FREKUENSI stores plain ints (format ``0``); IDs stay text.

    A float 7.0 under an integer format displays as `7` but round-trips through
    paste-values / CSV as `7.0`, which is not what the manual workbook shows.
    """
    out = tmp_path / "qlola_formats.xlsx"
    report = _run_qlola(qlola_raw_file, qlola_uker_reference_file, qlola_master_file, out)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Enriched_Data"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    id_col = headers.index("ID") + 1
    frek_col = headers.index("FREKUENSI") + 1

    for r in range(2, ws.max_row + 1):
        frek = ws.cell(row=r, column=frek_col)
        assert isinstance(frek.value, int) and not isinstance(frek.value, bool)
        assert frek.number_format == "0"
        # IDs are labels, not measures: written as text with no number format.
        id_cell = ws.cell(row=r, column=id_col)
        assert isinstance(id_cell.value, str)
        assert id_cell.number_format == "General"

    # U1's two rows (3 + 4) summed to exactly 7 — an int, not 7.0.
    frek_values = [ws.cell(row=r, column=frek_col).value for r in range(2, ws.max_row + 1)]
    assert 7 in frek_values


def test_qlola_summary_counts_are_plain_integers(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    """Crosstab counts are whole users; nothing renders as a formatted decimal."""
    out = tmp_path / "qlola_summary_formats.xlsx"
    report = _run_qlola(qlola_raw_file, qlola_uker_reference_file, qlola_master_file, out)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    header_row = next(r[0].row for r in ws.iter_rows() if r[0].value == "MAIN_CODE")
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None:
            break
        for c in range(2, 5):  # AKTIF / TIDAK AKTIF / Grand Total
            cell = ws.cell(row=r, column=c)
            assert isinstance(cell.value, int)
            assert cell.number_format == "0"


def test_briva_detail_exports_ceiled_plain_integers(
    timeseries_briva_file, reference_file, tmp_path
):
    """WF4 detail VOLUME_IDR cells are ceiled ints with plain ``0`` format."""
    out = tmp_path / "briva_detail_formats.xlsx"
    report = PipelineOrchestrator.execute(
        ProcessingConfig(
            raw_data_path=timeseries_briva_file,
            reference_data_path=reference_file,
            workflow_id="timeseries-fbi-briva",
            output_report_path=out,
        )
    )
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Enriched_Data"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    vol_col = headers.index("VOLUME_IDR") + 1
    cells = [ws.cell(row=r, column=vol_col) for r in range(2, ws.max_row + 1)]
    assert all(c.number_format == "0" for c in cells)
    assert all(isinstance(c.value, int) for c in cells)
    # Fractional cents in the fixture must ceil, not truncate.
    assert any(c.value % 1 == 0 for c in cells)


def test_healthy_qlola_run_shows_no_operator_diagnostic(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    """One unmapped ID out of five is a normal miss rate, not a join failure."""
    out = tmp_path / "qlola_healthy.xlsx"
    report = _run_qlola(qlola_raw_file, qlola_uker_reference_file, qlola_master_file, out)
    assert report.success is True, report.error_message
    assert report.unmapped_diagnostic is None
    assert unmapped_warning_text(report) is None


def test_failed_run_shows_no_unmapped_warning(tmp_path):
    """A hard failure surfaces its own error box; no stale unmapped warning."""
    report = PipelineOrchestrator.execute(
        ProcessingConfig(
            raw_data_path=tmp_path / "missing.csv",
            workflow_id="timeseries-active-user-qlola",
            output_report_path=tmp_path / "out.xlsx",
        )
    )
    assert report.success is False
    assert unmapped_warning_text(report) is None


# --------------------------------------------------------------------------- #
# Real-sample acceptance oracles (Path A' — skip if vendor samples absent)
# --------------------------------------------------------------------------- #
def test_qlola_real_sample_sheet1_oracle(qlola_sample_inputs, tmp_path):
    """Full crosstab parity with the QLOLA sample Sheet1 (3000/1820/4820; 7->130/81)."""
    out = tmp_path / "qlola_real.xlsx"
    config = ProcessingConfig(
        raw_data_path=qlola_sample_inputs["raw"],
        reference_data_path=qlola_sample_inputs["uker"],
        master_data_path=qlola_sample_inputs["master"],
        workflow_id="timeseries-active-user-qlola",
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    _, rows, grand_total = _read_crosstab_table(ws)

    # Per-MAIN_CODE parity (master ID->MAIN_CODE key, not UKER): code 7 -> 130/81.
    assert rows["7"]["AKTIF TRX >=5x"] == 130
    assert rows["7"]["TIDAK AKTIF <5x"] == 81
    # Sheet1 Grand Totals.
    assert grand_total["AKTIF TRX >=5x"] == 3000
    assert grand_total["TIDAK AKTIF <5x"] == 1820
    assert grand_total["Grand Total"] == 4820


def test_briva_real_sample_sheet1_oracle(briva_sample_inputs, tmp_path):
    """Grand Total parity with Excel ROUND/ceil of the Briva sample Sheet1 oracle."""
    out = tmp_path / "briva_real.xlsx"
    config = ProcessingConfig(
        raw_data_path=briva_sample_inputs["raw"],
        reference_data_path=briva_sample_inputs["uker"],
        workflow_id="timeseries-fbi-briva",
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    headers, _, grand_total = _read_summary_table(ws)
    assert "Sum of VOLUME_IDR" in headers
    assert grand_total["Sum of VOLUME_IDR"] == 2_481_517_421_604


def test_briva_real_sample_rows_are_mapped_not_unmapped(briva_sample_inputs, tmp_path):
    """The Grand Total alone is mapping-blind: assert the branches resolved too.

    The false-UNMAPPED collapse this change fixes produced a *correct* Grand
    Total on a single UNMAPPED row, so parity on the total proves nothing here.
    """
    out = tmp_path / "briva_real_mapping.xlsx"
    config = ProcessingConfig(
        raw_data_path=briva_sample_inputs["raw"],
        reference_data_path=briva_sample_inputs["uker"],
        workflow_id="timeseries-fbi-briva",
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message
    assert report.unmapped_records_count == 0

    _, rows, _ = _read_summary_table(load_workbook(out)["Summary_Report"])
    assert len(rows) > 1
    assert "UNMAPPED" not in rows
    # MAIN_BRANCH must carry real branch names, not the UNMAPPED sentinel.
    assert all(r["MAIN_BRANCH"] not in (None, "", "UNMAPPED") for r in rows.values())


def test_qlola_real_sample_rows_are_mapped_not_unmapped(qlola_sample_inputs, tmp_path):
    """Companion mapping assertion for the Qlola crosstab (see the Briva twin)."""
    out = tmp_path / "qlola_real_mapping.xlsx"
    config = ProcessingConfig(
        raw_data_path=qlola_sample_inputs["raw"],
        reference_data_path=qlola_sample_inputs["uker"],
        master_data_path=qlola_sample_inputs["master"],
        workflow_id="timeseries-active-user-qlola",
        output_report_path=out,
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message
    assert report.unmapped_records_count == 0

    _, rows, grand_total = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    assert len(rows) > 1
    assert "UNMAPPED" not in rows
    # Column sums reconcile with the Grand Total row.
    assert sum(r["AKTIF TRX >=5x"] for r in rows.values()) == grand_total["AKTIF TRX >=5x"]
    assert (
        sum(r["TIDAK AKTIF <5x"] for r in rows.values())
        == grand_total["TIDAK AKTIF <5x"]
    )


# --------------------------------------------------------------------------- #
# Live acceptance (operator Downloads folder — skipped in CI)
# --------------------------------------------------------------------------- #
_LIVE_QLola_JUL24_RAW = Path(
    r"C:\Users\Draven Chen\Downloads"
    r"\1785047437_REPORT_SUMMARY_MONTHLY_2026-07-24_QLOLA_ALL.csv"
)
_LIVE_QLola_JUL24_UKER = Path(
    r"C:\Users\Draven Chen\Downloads"
    r"\Mapping GL FBI Vol Uker (BRIEFX) Update BO BATU.xlsx"
)
_LIVE_QLola_JUL24_MASTER = Path(
    r"C:\Users\Draven Chen\Downloads\MASTER DATA STATIS MALANG JUN (3).xlsx"
)


@pytest.mark.skipif(
    not _LIVE_QLola_JUL24_RAW.exists()
    or not _LIVE_QLola_JUL24_UKER.exists()
    or not _LIVE_QLola_JUL24_MASTER.exists(),
    reason="Jul-24 Qlola acceptance files not in Downloads",
)
def test_qlola_jul24_live_acceptance_oracle(tmp_path):
    """Parity with manual Sheet1 pivot for the Jul-24 QLOLA_ALL operator run."""
    out = tmp_path / "qlola_jul24_live.xlsx"
    report = _run_qlola(
        _LIVE_QLola_JUL24_RAW, _LIVE_QLola_JUL24_UKER, _LIVE_QLola_JUL24_MASTER, out
    )
    assert report.success is True, report.error_message
    assert report.unmapped_records_count == 372
    assert report.unmapped_diagnostic is None

    _, rows, grand_total = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    assert len(rows) == 25
    assert rows["7"]["AKTIF TRX >=5x"] == 139
    assert rows["7"]["TIDAK AKTIF <5x"] == 75
    # Mapped-only Grand Totals (dashboard contract: exclude #N/A / UNMAPPED).
    assert grand_total["AKTIF TRX >=5x"] == 3039
    assert grand_total["TIDAK AKTIF <5x"] == 1626
    assert grand_total["Grand Total"] == 4665


# --------------------------------------------------------------------------- #
# Unified FILTERS block: empty = automatic, written = keep-list override
#
# `_gui_filters` reproduces exactly what `launch_gui`'s run() builds from the
# three FILTERS boxes, so these exercise the operator's real contract rather
# than a hand-written config that only resembles it.
# --------------------------------------------------------------------------- #
def _gui_filters(segment: str = "", source: str = "", kw: str = "") -> dict:
    """Config kwargs for the given SEGMENT / SOURCE / KW box contents."""
    seg = parse_filter_input(segment)
    src = parse_filter_input(source)
    return {
        "segmen_filter": seg,
        # A typed keep-list replaces that dimension's default policy for the run.
        "segmen_exclude": None if seg is None else [],
        "source_include": src,
        "source_exclude": None if src is None else [],
        "kw_include": parse_filter_input(kw),
    }


def test_empty_boxes_send_none_to_the_workflow():
    """The whole revamp rests on this: empty must not mean "filter nothing"."""
    assert _gui_filters() == {
        "segmen_filter": None,
        "segmen_exclude": None,
        "source_include": None,
        "source_exclude": None,
        "kw_include": None,
    }


def test_empty_segment_box_still_drops_korporasi(
    report_data_statis_bin_file, reference_file, tmp_path
):
    """Data Statis' baked KORPORASI exclusion is what an empty box runs."""
    out = tmp_path / "statis_empty_boxes.xlsx"
    report = PipelineOrchestrator.execute(
        _statis_config(report_data_statis_bin_file, reference_file, out, **_gui_filters())
    )
    assert report.success is True, report.error_message

    rows = _statis_summary_rows(load_workbook(out)["Summary_Report"])
    # P4/P5 (KORPORASI) stay out, exactly as with no filter arguments at all.
    assert rows[("MC10", "Jakarta Pusat")].value == 3


def test_empty_source_box_still_drops_cms(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    """Qlola's baked CMS exclusion is what an empty box runs."""
    out = tmp_path / "qlola_empty_boxes.xlsx"
    config = ProcessingConfig(
        raw_data_path=qlola_raw_file,
        reference_data_path=qlola_uker_reference_file,
        master_data_path=qlola_master_file,
        workflow_id="timeseries-active-user-qlola",
        output_report_path=out,
        **_gui_filters(),
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    _, rows, grand_total = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    # Identical to the untouched-default run: U6 (CMS-only) never counts.
    assert grand_total["Grand Total"] == 4
    assert rows["9"]["AKTIF TRX >=5x"] == 1


def test_empty_segment_box_still_keeps_only_nonwholesale(tmp_path):
    """Briva's default is an *inclusion*, and an empty box must run it too."""
    strategy = get_strategy(WorkflowId.TIMESERIES_FBI_BRIVA)
    config = ProcessingConfig(
        raw_data_path=tmp_path / "raw.csv",
        workflow_id="timeseries-fbi-briva",
        output_report_path=tmp_path / "out.xlsx",
        **_gui_filters(),
    )
    assert strategy.resolve_segment_filter(config) == "NONWHOLESALE"


def test_written_segment_box_replaces_the_baked_exclusion(
    report_data_statis_bin_file, reference_file, tmp_path
):
    """Typing the segment the report drops by default must keep only that one."""
    out = tmp_path / "statis_korporasi_only.xlsx"
    report = PipelineOrchestrator.execute(
        _statis_config(
            report_data_statis_bin_file,
            reference_file,
            out,
            **_gui_filters(segment="KORPORASI"),
        )
    )
    assert report.success is True, report.error_message

    rows = _statis_summary_rows(load_workbook(out)["Summary_Report"])
    # P4 + P5: the default KORPORASI exclusion was replaced, not AND-ed.
    assert rows[("MC10", "Jakarta Pusat")].value == 2


def test_written_segment_box_accepts_a_comma_separated_keep_list(
    report_data_statis_bin_file, reference_file, tmp_path
):
    out = tmp_path / "statis_two_segments.xlsx"
    report = PipelineOrchestrator.execute(
        _statis_config(
            report_data_statis_bin_file,
            reference_file,
            out,
            **_gui_filters(segment="Consumer, KORPORASI"),
        )
    )
    assert report.success is True, report.error_message

    rows = _statis_summary_rows(load_workbook(out)["Summary_Report"])
    # P1 (Consumer) + P4 + P5 (KORPORASI); the Micro/SME/blank rows drop out.
    assert rows[("MC10", "Jakarta Pusat")].value == 3


def test_written_source_box_replaces_the_qlola_cms_exclusion(
    qlola_raw_file, qlola_uker_reference_file, qlola_master_file, tmp_path
):
    """Typing CMS keeps only CMS - the exact inverse of the baked default."""
    out = tmp_path / "qlola_cms_only.xlsx"
    config = ProcessingConfig(
        raw_data_path=qlola_raw_file,
        reference_data_path=qlola_uker_reference_file,
        master_data_path=qlola_master_file,
        workflow_id="timeseries-active-user-qlola",
        output_report_path=out,
        **_gui_filters(source="CMS"),
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    _, rows, grand_total = _read_crosstab_table(load_workbook(out)["Summary_Report"])
    # Only U1's CMS row (master 7) and U6 (master 9) survive, both active.
    assert grand_total["Grand Total"] == 2
    assert rows["7"]["AKTIF TRX >=5x"] == 1
    assert rows["9"]["AKTIF TRX >=5x"] == 1


def test_typed_filters_are_a_no_op_on_an_extract_without_those_columns(
    rincian_portal_bg_file, tmp_path
):
    """MAINBR/MBNAME/AMOUNT only: all three filters must pass every row through."""
    out = tmp_path / "portal_bg_all_filters.xlsx"
    config = ProcessingConfig(
        raw_data_path=rincian_portal_bg_file,
        workflow_id="rincian-portal-bg",
        output_report_path=out,
        **_gui_filters(segment="Consumer", source="CMS", kw="KW01"),
    )
    report = PipelineOrchestrator.execute(config)
    assert report.success is True, report.error_message

    ws = load_workbook(out)["Summary_Report"]
    rows = {
        (r[0].value, r[1].value): r[2].value
        for r in ws.iter_rows()
        if r[0].value in {"B01", "B02"}
    }
    assert rows[("B01", "Branch One")] == 3000
    assert rows[("B02", "Branch Two")] == 500


def test_giro_typed_filters_are_a_no_op_without_those_columns(
    giro_monthly_file, giro_master_file, tmp_path
):
    """Report Giro spec: a typed filter cannot fail a NO REK + SALDO IDR run."""
    out = tmp_path / "giro_all_filters_noop.xlsx"
    report = _run_giro(
        giro_monthly_file,
        giro_master_file,
        out,
        **_gui_filters(segment="RITEL", source="BRANCH", kw="KW01"),
    )
    assert report.success is True, report.error_message

    ws = load_workbook(out)["DAFTAR GIRO"]
    assert ws.cell(GIRO_ROW_MATCHED, GIRO_SALDO_UPDATE_COL).value == 2500
    assert ws.cell(GIRO_ROW_PLAIN, GIRO_SALDO_UPDATE_COL).value == 7500


def test_giro_typed_segment_box_keeps_only_that_segment(
    giro_monthly_file_with_segmen, giro_master_file, tmp_path
):
    """Where the monthly source does carry SEGMEN, the box must really filter."""
    out = tmp_path / "giro_gui_segment.xlsx"
    report = _run_giro(
        giro_monthly_file_with_segmen,
        giro_master_file,
        out,
        **_gui_filters(segment="RITEL"),
    )
    assert report.success is True, report.error_message

    ws = load_workbook(out)["DAFTAR GIRO"]
    assert ws.cell(GIRO_ROW_MATCHED, GIRO_SALDO_UPDATE_COL).value is None  # KORPORASI
    assert ws.cell(GIRO_ROW_PLAIN, GIRO_SALDO_UPDATE_COL).value == 7500


# --------------------------------------------------------------------------- #
# KW dimension (AggregationEngine)
# --------------------------------------------------------------------------- #
def _kw_frame(**columns) -> pd.DataFrame:
    base = {
        "MAIN_CODE": ["A", "B"],
        "MAIN_BRANCH": ["Branch A", "Branch B"],
        "VOLUME_IN_IDR": [100.0, 200.0],
    }
    base.update(columns)
    return pd.DataFrame(base)


def test_kw_filter_keeps_only_the_requested_kw():
    result = AggregationEngine(kw_include=["KW01"]).aggregate(
        _kw_frame(KW=["KW01", "KW02"])
    )
    assert result.total_volume_idr == 100.0
    assert result.branch_count == 1


def test_kw_filter_accepts_a_multi_value_keep_list():
    result = AggregationEngine(kw_include=["KW01", "kw02"]).aggregate(
        _kw_frame(KW=["KW01", "KW02"])
    )
    assert result.total_volume_idr == 300.0


@pytest.mark.parametrize("header", ["KW", "PRODUCT", "GROUP_PRODUCT", "KAWIL"])
def test_kw_filter_resolves_each_header_alias(header):
    """The stakeholder's KW may be spelled several ways in the extract."""
    result = AggregationEngine(kw_include=["KW01"]).aggregate(
        _kw_frame(**{header: ["KW01", "KW02"]})
    )
    assert result.total_volume_idr == 100.0


def test_kw_filter_is_a_no_op_without_a_kw_column():
    """Missing-column pass-through: every row survives and nothing raises."""
    result = AggregationEngine(kw_include=["KW01"]).aggregate(_kw_frame())
    assert result.total_volume_idr == 300.0
    assert result.branch_count == 2


def test_empty_kw_keep_list_keeps_every_row():
    result = AggregationEngine(kw_include=[]).aggregate(_kw_frame(KW=["KW01", "KW02"]))
    assert result.total_volume_idr == 300.0


def test_all_filter_dimensions_are_a_no_op_on_a_bare_frame():
    """SEGMENT + SOURCE + KW typed against a frame that has none of them."""
    engine = AggregationEngine(
        segment_filter=["Consumer"],
        exclude_segmen=["KORPORASI"],
        kw_include=["KW01"],
    )
    result = engine.aggregate(_kw_frame())
    assert result.total_volume_idr == 300.0
    assert result.branch_count == 2


def test_kw_override_is_gated_by_the_workflow_flag(tmp_path):
    """`has_kw_filter` gates the runtime override exactly like the other flags."""
    config = _flagless_config(tmp_path, kw_include=["KW01"])
    assert get_strategy(WorkflowId.REPORT_GIRO).resolve_kw_include(config) == ["KW01"]
    assert _FlaglessStrategy().resolve_kw_include(config) == []
