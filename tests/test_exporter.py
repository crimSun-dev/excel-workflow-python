"""Phase 2 integration tests: Excel export & formatting (TDD Section 4)."""

from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.exporter import ExcelReportExporter, sanitize_cell_value


@pytest.fixture
def summary_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "MAIN_CODE": ["MC10", "MC20"],
            "MAIN_BRANCH": ["Jakarta Pusat", "Surabaya"],
            "VOLUME_IN_IDR": [1_500_000_000_000.0, 500_000_000.50],
        }
    )


@pytest.fixture
def enriched_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "KODE_UKER": ["0001", "0002"],
            "SEGMEN": ["Wholesale", "Corporate"],
            "VOLUME_IN_IDR": [1_500_000_000_000.0, 500_000_000.50],
            "MAIN_CODE": ["MC10", "MC20"],
            "MAIN_BRANCH": ["Jakarta Pusat", "Surabaya"],
        }
    )


def test_exporter_generates_valid_xlsx_workbook(summary_df, enriched_df, tmp_path):
    out = tmp_path / "report.xlsx"
    ExcelReportExporter().export(summary_df, enriched_df, out)
    assert out.exists()
    wb = load_workbook(out)  # readable by openpyxl
    assert wb is not None


def test_exporter_creates_summary_and_enriched_sheets(summary_df, enriched_df, tmp_path):
    out = tmp_path / "report.xlsx"
    ExcelReportExporter().export(summary_df, enriched_df, out)
    wb = load_workbook(out)
    assert "Summary_Report" in wb.sheetnames
    assert "Enriched_Data" in wb.sheetnames


def test_exporter_applies_explicit_number_format(summary_df, enriched_df, tmp_path):
    out = tmp_path / "report.xlsx"
    ExcelReportExporter().export(summary_df, enriched_df, out)
    ws = load_workbook(out)["Summary_Report"]
    # Find a cell holding the large volume value and check its format.
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value == 1_500_000_000_000.0:
                assert cell.number_format == "0"
                found = True
    assert found, "Volume value cell not found in summary sheet"


def test_exporter_no_scientific_notation_large_numbers(summary_df, enriched_df, tmp_path):
    out = tmp_path / "report.xlsx"
    ExcelReportExporter().export(summary_df, enriched_df, out)
    ws = load_workbook(out)["Summary_Report"]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                # Value stored as a true float, not a string like "1.5E+12".
                assert not isinstance(cell.value, str)


def test_exporter_adds_grand_total_row(summary_df, enriched_df, tmp_path):
    out = tmp_path / "report.xlsx"
    ExcelReportExporter().export(summary_df, enriched_df, out)
    ws = load_workbook(out)["Summary_Report"]
    labels = [c.value for row in ws.iter_rows() for c in row if c.value == "Grand Total"]
    assert "Grand Total" in labels
    # Grand Total value should equal the sum of the volume column (plain ints).
    total = int(round(summary_df["VOLUME_IN_IDR"].sum()))
    values = [
        c.value for row in ws.iter_rows() for c in row if c.value == total
    ]
    assert total in values


def test_exporter_auto_adjusts_column_widths(summary_df, enriched_df, tmp_path):
    out = tmp_path / "report.xlsx"
    ExcelReportExporter().export(summary_df, enriched_df, out)
    ws = load_workbook(out)["Summary_Report"]
    # At least one column width was explicitly set (> default of None/8.43).
    widths = [dim.width for dim in ws.column_dimensions.values() if dim.width]
    assert widths and max(widths) > 10


# --------------------------------------------------------------------------- #
# Illegal control characters in raw extract text
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("MUHAMMAD ZAPA AL AS\x00", "MUHAMMAD ZAPA AL AS"),
        ("A\x01B\x1fC", "ABC"),
        ("already clean", "already clean"),
        # Tab / newline / carriage return are legal in xlsx and must survive.
        ("line1\nline2\tend\r", "line1\nline2\tend\r"),
    ],
)
def test_sanitize_cell_value_strips_only_illegal_characters(raw, expected):
    assert sanitize_cell_value(raw) == expected


@pytest.mark.parametrize("value", [None, 42, 3.5, True])
def test_sanitize_cell_value_passes_non_strings_through_unchanged(value):
    assert sanitize_cell_value(value) is value


def test_exporter_writes_rows_containing_control_characters(summary_df, tmp_path):
    """Regression: one stray NUL in a customer name aborted the whole export."""
    enriched = pd.DataFrame(
        {
            "KODE_UKER": ["0001"],
            "NAMA_REK_SIMPANAN": ["MUHAMMAD ZAPA AL AS\x00"],
            "VOLUME_IN_IDR": [1000.0],
            "MAIN_CODE": ["MC10"],
            "MAIN_BRANCH": ["Jakarta Pusat"],
        }
    )
    out = tmp_path / "control_chars.xlsx"
    ExcelReportExporter().export(summary_df, enriched, out)

    ws = load_workbook(out)["Enriched_Data"]
    names = [c.value for row in ws.iter_rows(min_row=2) for c in row]
    assert "MUHAMMAD ZAPA AL AS" in names


def test_exporter_cleans_control_characters_in_headers_and_summary(tmp_path):
    """The junk byte can land in a header or a branch name, not just a detail row."""
    summary = pd.DataFrame(
        {
            "MAIN_CODE": ["MC10"],
            "MAIN_BRANCH": ["Surabaya\x00"],
            "VOLUME_IN_IDR": [1000.0],
        }
    )
    enriched = pd.DataFrame({"NAMA\x00": ["value\x01"], "VOLUME_IN_IDR": [1000.0]})
    out = tmp_path / "control_chars_headers.xlsx"
    ExcelReportExporter().export(summary, enriched, out)

    workbook = load_workbook(out)
    summary_values = [
        c.value for row in workbook["Summary_Report"].iter_rows() for c in row
    ]
    assert "Surabaya" in summary_values
    detail = workbook["Enriched_Data"]
    assert detail.cell(row=1, column=1).value == "NAMA"
    assert detail.cell(row=2, column=1).value == "value"


def test_exporter_writes_missing_detail_cells_as_blank(summary_df, tmp_path):
    enriched = pd.DataFrame(
        {
            "KODE_UKER": ["0001"],
            "SEGMEN": [pd.NA],
            "VOLUME_IN_IDR": [1000.0],
        }
    )
    out = tmp_path / "missing_detail.xlsx"
    ExcelReportExporter().export(summary_df, enriched, out)
    ws = load_workbook(out)["Enriched_Data"]
    assert ws.cell(row=2, column=2).value is None
